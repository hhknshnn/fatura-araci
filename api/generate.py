from http.server import BaseHTTPRequestHandler
import json
import base64
import io
import os
import re
import traceback
import pdfplumber

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# ── SABITLER ──────────────────────────────────────────────────────────────────
DARK_BLUE  = '1F3864'
MID_BLUE   = '2F5496'
LIGHT_BLUE = 'D6E4F0'
GOLD       = 'C9A84C'
LIGHT_GRAY = 'F2F2F2'
TL_FMT     = '₺#,##0.00'

EXCEPTION_SKUS = {
    '1SPOCA0029197.': 0.01,
    '1SPOCA0030197.': 0.01,
    '1SSARF1507139.': 0.01,
    '1SPOCA0027197.': 0.01,
    '1SPOCA0028169.': 0.01,
    '1SPOCA0030169.': 0.01,
    '1SPOCA0028197.': 0.01,
}

INV_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('UNIT',              'Birim Cinsi (1)'),
    ('QTY',               'Miktar'),
    ('UNIT PRICE',        'Fiyat'),
    ('TOTAL AMOUNT TRY',  '__CALC__'),
    ('HS CODE',           'GTİP'),
    ('MATERIAL',          'MATERYAL -EN'),
    ('ITEM NAME-Serb',    'Ürün Açıklaması XS'),
    ('COLOR SERB',        'Renk Açıkmalası XS'),
    ('MATERIAL SERB',     'MATERYAL -XS'),
    ('DIMENSION',         'EBAT Açıklama'),
]

PL_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('UNIT',              'Birim Cinsi (1)'),
    ('QTY',               'Miktar'),
    ('GROSS WEIGHT',      '__BRUT__'),
    ('NET WEIGHT',        '__NET__'),
]

BA_INV_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('UNIT PRICE',        'Fiyat (D)'),
    ('TOTAL AMOUNT TRY',  'Net Tutar (D)'),
    ('HS CODE',           'GTİP'),
    ('MATERIAL',          'MATERYAL -EN'),
    ('COLOR',             'Renk Açıkmalası EN'),
    ('DIMENSION',         'EBAT Açıklama'),
]

BA_PL_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('GROSS WEIGHT',      '__BRUT__'),
    ('NET WEIGHT',        '__NET__'),
]

GE_INV_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('HS CODE',           'GTİP'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('UNIT PRICE',        'Fiyat'),
    ('TOTAL AMOUNT TRY',  '__CALC__'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('BARCODE',           'Asorti Barkodu'),
    ('MATERIAL',          'MATERYAL -EN'),
    ('DIMENSION',         'EBAT Açıklama'),
]

GE_PL_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('HS CODE',           'GTİP'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('GROSS WEIGHT',      '__BRUT__'),
    ('NET WEIGHT',        '__NET__'),
]

KO_INV_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('HS CODE',           'GTİP'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('UNIT PRICE',        '__EUR__'),
    ('TOTAL AMOUNT EUR',  '__EUR_CALC__'),
    ('MATERIAL',          'MATERYAL -EN'),
    ('COLOR',             'Renk Açıkmalası EN'),
    ('DIMENSION',         'EBAT Açıklama'),
]

KO_PL_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('HS CODE',           'GTİP'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('GROSS WEIGHT',      '__BRUT__'),
    ('NET WEIGHT',        '__NET__'),
]

BE_INV_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('HS CODE',           'GTİP'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('UNIT PRICE',        '__EUR__'),
    ('TOTAL AMOUNT EUR',  '__EUR_CALC__'),
    ('MATERIAL',          'MATERYAL -EN'),
    ('COLOR',             'Renk Açıkmalası EN'),
    ('DIMENSION',         'EBAT Açıklama'),
]

BE_PL_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('HS CODE',           'GTİP'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('GROSS WEIGHT',      '__BRUT__'),
    ('NET WEIGHT',        '__NET__'),
]

KZ_INV_COLS = [
    ('COUNTRY OF ORIGIN',    'MENŞEİ -EN'),
    ('СТРАНА ПРОИСХОЖДЕНИЯ', 'MENŞEİ -RU'),
    ('Master Carton Code',   'Asorti Barkodu'),
    ('ITEM CODE',            'SKU'),
    ('COLOR NAME',           'Renk Açıkmalası EN'),
    ('HS CODE',              'GTİP'),
    ('ITEM DESCRIPTION - EN','ALT GRUBU -EN'),
    ('ITEM NAME - EN',       'Ürün Açıklaması EN'),
    ('описание товаров',     'Ürün Açıklaması RU'),
    ('UNIT',                 'Miktar'),
    ('UNIT PRICE',           'Fiyat'),
    ('TOTAL AMOUNT',         '__CALC__'),
    ('MATERIAL',             'MATERYAL -EN'),
    ('МАТЕРИАЛ',             'MATERYAL -RU'),
    ('ALT GRUBU Açıklama',   'ALT GRUBU Açıklama'),
    ('DIMENSION',            'EBAT Açıklama'),
]

KZ_PL_COLS = [
    ('COUNTRY OF ORIGIN',    'MENŞEİ -EN'),
    ('СТРАНА ПРОИСХОЖДЕНИЯ', 'MENŞEİ -RU'),
    ('Master Carton Code',   'Asorti Barkodu'),
    ('ITEM CODE',            'SKU'),
    ('COLOR NAME',           'Renk Açıkmalası EN'),
    ('HS CODE',              'GTİP'),
    ('ITEM DESCRIPTION - EN','ALT GRUBU -EN'),
    ('ITEM NAME - EN',       'Ürün Açıklaması EN'),
    ('описание товаров',     'Ürün Açıklaması RU'),
    ('UNIT',                 'Miktar'),
    ('GROSS WEIGHT',         '__BRUT__'),
    ('NET WEIGHT',           '__NET__'),
]

DE_INV_COLS = BE_INV_COLS
DE_PL_COLS  = BE_PL_COLS
NL_INV_COLS = BE_INV_COLS
NL_PL_COLS  = BE_PL_COLS

RU_INV_COLS = KZ_INV_COLS  # Rusya — Kazakistan ile aynı yapı
RU_PL_COLS  = KZ_PL_COLS

UZ_INV_COLS = KZ_INV_COLS  # Özbekistan — Kazakistan ile aynı yapı
UZ_PL_COLS  = KZ_PL_COLS

# Genel USD ülkeleri — Irak, Libya, Liberya, Lübnan
GENEL_INV_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('HS CODE',           'GTİP'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('UNIT PRICE',        '__USD__'),
    ('TOTAL AMOUNT USD',  '__USD_CALC__'),
    ('MATERIAL',          'MATERYAL -EN'),
    ('COLOR',             'Renk Açıkmalası EN'),
    ('DIMENSION',         'EBAT Açıklama'),
]

GENEL_PL_COLS = [
    ('COUNTRY OF ORIGIN', 'MENŞEİ -EN'),
    ('MASTER ITEM CODE',  'Asorti Barkodu'),
    ('ITEM CODE',         'SKU'),
    ('HS CODE',           'GTİP'),
    ('ITEM DESCRIPTION',  'ALT GRUBU -EN'),
    ('ITEM NAME',         'Ürün Açıklaması EN'),
    ('QTY',               'Miktar'),
    ('GROSS WEIGHT',      '__BRUT__'),
    ('NET WEIGHT',        '__NET__'),
]

def brd(c='BFBFBF'):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, r, col, val, bg=DARK_BLUE, fg='FFFFFF', bold=True, align='center', size=9):
    c = ws.cell(row=r, column=col, value=val)
    c.font = Font(name='Arial', bold=bold, color=fg, size=size)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = brd()
    return c

def dat(ws, r, col, val, bg='FFFFFF', bold=False, align='left', fmt=None):
    c = ws.cell(row=r, column=col, value=val)
    c.font = Font(name='Arial', bold=bold, color='000000', size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = brd()
    if fmt: c.number_format = fmt
    return c

def parse_num(v):
    if v is None or v == '': return 0.0
    if isinstance(v, (int, float)):
        return float(v) if str(v) not in ['nan','inf'] else 0.0
    s = str(v).strip().replace(' ','').replace('\u00a0','')
    if '.' in s and ',' in s:
        s = s.replace('.','').replace(',','.')
    elif ',' in s:
        s = s.replace(',','.')
    try: return float(s)
    except: return 0.0

def _normalize_pdf_text(text):
    return re.sub(r'\s+', ' ', (text or '').replace('\u00a0', ' ')).strip()

def _extract_pdf_amount(text, patterns):
    def _parse_pdf_amount(value):
        s = str(value).strip().replace(' ', '').replace('\u00a0', '')
        if '.' in s and ',' in s:
            if s.rfind(',') > s.rfind('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            s = s.replace(',', '.')
        try:
            return float(s)
        except Exception:
            return parse_num(value)
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            return _parse_pdf_amount(m.group(1))
    return 0.0

def _extract_pdf_packages(text):
    patterns = [
        r'[*\-]?\s*KAP\s+ADET[İI]\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'[*\-]?\s*KAP\s+SAYISI\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'[*\-]?\s*KAP\s+ADEDI\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'[*\-]?\s*KAP\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'\bPACKAGES?\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'\bCOLL[Iİ]\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
    ]
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return ''

def parse_pdf(pdf_bytes):
    result = {'navlun': 0.0, 'sigorta': 0.0, 'kap': ''}
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            page_texts = [_normalize_pdf_text(page.extract_text() or '') for page in pdf.pages]
            text = ' '.join(part for part in page_texts if part).strip()
            if not text:
                return result
            result['navlun'] = _extract_pdf_amount(text, [
                r'\bNAVLUN\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
                r'\bFREIGHT\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
            ])
            result['sigorta'] = _extract_pdf_amount(text, [
                r'S[İI]G(?:ORTA)?\.?\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
                r'\bINSURANCE\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
            ])
            result['kap'] = _extract_pdf_packages(text)
    except Exception:
        pass
    return result

def calculate_weights(df, grup_kilolari, hedef_brut, exception_skus):
    ham_list = []
    for _, row in df.iterrows():
        sku    = str(row.get('SKU','')).strip()
        grup   = str(row.get('ÜRÜN ARA GRUBU','')).strip()
        ag     = parse_num(row.get('Ürün Ağırlığı (KG)',0))
        miktar = parse_num(row.get('Miktar',0))
        if sku in exception_skus:
            kg = parse_num(exception_skus[sku])
        elif ag > 0:
            kg = ag
        else:
            kg = parse_num(grup_kilolari.get(grup,0))
        ham_list.append(kg * miktar)
    ham_toplam = sum(ham_list)
    if ham_toplam <= 0:
        return [0.0]*len(ham_list), [0.0]*len(ham_list)
    carpan = hedef_brut / ham_toplam

    brut_list = []
    toplam_yuvarlanmis = 0.0
    for i, h in enumerate(ham_list):
        if i < len(ham_list) - 1:
            val = round(h * carpan, 2)
            brut_list.append(val)
            toplam_yuvarlanmis += val
        else:
            brut_list.append(round(hedef_brut - toplam_yuvarlanmis, 2))

    hedef_net_serbest = round(hedef_brut * 0.9, 2)
    net_list = []
    toplam_net = 0.0
    for i, b in enumerate(brut_list):
        if i < len(brut_list) - 1:
            val = round(b * 0.9, 2)
            net_list.append(val)
            toplam_net += val
        else:
            net_list.append(round(hedef_net_serbest - toplam_net, 2))
    return brut_list, net_list

def _sku_grupla(df):
    """SKU bazında gruplandırma — tüm ülkeler için ortak."""
    agg_dict = {col: 'first' for col in df.columns if col != 'SKU'}
    agg_dict['Miktar'] = 'sum'
    return df.groupby('SKU', sort=False).agg(agg_dict).reset_index()

def generate_master_excel(df_original, brut_list, net_list):
    """
    Orijinal master Excel'e BRÜT ve NET sütunları ekler.
    brut_list/net_list orijinal df ile aynı uzunlukta olmalı.
    BRÜT / Miktar → Ürün Ağırlığı (KG) sütununa yazılır.
    """
    df = df_original.copy()
    df['BRÜT'] = brut_list
    df['NET']  = net_list

    def calc_ag(row):
        miktar = parse_num(row.get('Miktar', 0))
        brut   = parse_num(row.get('BRÜT', 0))
        if miktar > 0:
            return round(brut / miktar, 6)
        return 0.0

    df['Ürün Ağırlığı (KG)'] = df.apply(calc_ag, axis=1)

    cols = list(df.columns)
    ag_idx = cols.index('Ürün Ağırlığı (KG)')
    cols.remove('BRÜT')
    cols.remove('NET')
    cols.insert(ag_idx + 1, 'BRÜT')
    cols.insert(ag_idx + 2, 'NET')
    df = df[cols]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    buf.seek(0)
    return buf.getvalue()

def build_header(ws, sheet_title, fatura_no, fatura_date, musteri, musteri_adres, col_count, logo_bytes=None, pdf_fields=None, destination='SERBIA', incoterm='CIP', info_start_col=None):
    last_col = get_column_letter(col_count)
    col_widths = {'A':16,'B':14,'C':14,'D':18,'E':33,'F':6,'G':7,'H':26,'I':22,'J':13,'K':12,'L':16,'M':9,'N':9,'O':10}
    for col, w in col_widths.items():
        if column_index_from_string(col) <= col_count:
            ws.column_dimensions[col].width = w

    if info_start_col is None:
        info_start_col = min(8, max(1, col_count - 1))
    info_value_col = min(col_count, info_start_col + 1)
    title_end_col = get_column_letter(info_value_col)
    hc = get_column_letter(max(1, info_start_col - 1))

    ws.row_dimensions[1].height = 27.0
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].fill = PatternFill('solid', fgColor='FFFFFF')
    if logo_bytes:
        try:
            from PIL import Image as PILImage
            import numpy as np
            pil_img = PILImage.open(io.BytesIO(logo_bytes))
            arr = np.array(pil_img.convert('RGB'))
            mask = (arr < 240).any(axis=2)
            rows = np.any(mask, axis=1)
            cols = np.any(mask, axis=0)
            if rows.any() and cols.any():
                rmin,rmax = np.where(rows)[0][[0,-1]]
                cmin,cmax = np.where(cols)[0][[0,-1]]
                pad = 10
                pil_img = pil_img.crop((
                    max(0,cmin-pad), max(0,rmin-pad),
                    min(arr.shape[1],cmax+pad), min(arr.shape[0],rmax+pad)
                ))
            logo_buf = io.BytesIO()
            pil_img.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            img = XLImage(logo_buf)
            img.width, img.height = 240, 22
            ws.add_image(img, 'E1')
        except Exception:
            pass

    ws.row_dimensions[2].height = 28.0
    ws.merge_cells(f'A2:{title_end_col}2')
    c = ws['A2']
    c.value = sheet_title
    c.font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    for col_idx in range(info_value_col + 1, column_index_from_string(last_col)+1):
        ws.cell(row=2, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    for i in range(3, 9):
        ws.row_dimensions[i].height = 22

    def info_label(r, txt):
        c = ws.cell(row=r, column=info_start_col, value=txt)
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill = PatternFill('solid', fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = brd()

    def info_val(r, val, bold=False):
        c = ws.cell(row=r, column=info_value_col, value=val)
        c.font = Font(name='Arial', bold=bold, color='000000', size=9)
        c.fill = PatternFill('solid', fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = brd()

    ws.merge_cells(f'A3:{hc}3')
    hdr(ws, 3, 1, 'EXPORTER :', bg=MID_BLUE, align='left')
    info_label(3, 'INVOICE DATE :')
    info_val(3, str(fatura_date), bold=True)

    ws.row_dimensions[4].height = 32
    ws.merge_cells(f'A4:{hc}5')
    c = ws['A4']
    c.value = ('DEHA MAGAZACILIK EV TEKSTILI URUNLERI SAN. VE TIC. A.S.\n'
               'Mecidiyeköy Mah. Oğuz Sok Rönesans Biz İş Merkezi No:4/14 K:4 34387 Şişli/İstanbul')
    c.font = Font(name='Arial', size=8, color='000000')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill = PatternFill('solid', fgColor=LIGHT_GRAY)
    c.border = brd()
    info_label(4, 'INVOICE NO :')
    info_val(4, fatura_no, bold=True)
    info_label(5, 'PACKAGES :')
    info_val(5, str(pdf_fields.get('kap','')) if pdf_fields else '')

    ws.merge_cells(f'A6:{hc}6')
    hdr(ws, 6, 1, 'IMPORTER :', bg=MID_BLUE, align='left')
    info_label(6, 'DESTINATION :')
    info_val(6, destination, bold=True)

    ws.row_dimensions[7].height = 32
    ws.merge_cells(f'A7:{hc}8')
    c = ws['A7']
    c.value = f'{musteri}\n{musteri_adres}'
    c.font = Font(name='Arial', size=9, color='000000')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill = PatternFill('solid', fgColor=LIGHT_GRAY)
    c.border = brd()
    info_label(7, 'COUNTRY OF EXPORTATION :')
    info_val(7, 'TURKEY')
    info_label(8, 'INCOTERM :')
    info_val(8, incoterm)

def build_footer(ws, footer_start, col_count):
    pass

def set_print(ws, print_area):
    ws.print_area = print_area
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75, header=0.3, footer=0.3)
    ws.print_title_rows = '1:2'

def apply_ba_template_header(ws, sheet_title, fatura_no, fatura_date, musteri, musteri_adres,
                             destination, incoterm, packages=''):
    ws['A2'] = sheet_title
    ws['H3'] = str(fatura_date)
    ws['H4'] = fatura_no
    ws['H5'] = packages
    ws['H6'] = destination
    ws['H7'] = 'TURKEY'
    ws['H8'] = incoterm
    ws['A7'] = f'{musteri}\n{musteri_adres}'

def apply_ge_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['H3'] = str(fatura_date)
    ws['H4'] = fatura_no
    ws['H5'] = packages

def apply_ko_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages

def apply_mk_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages

def apply_be_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages

def apply_de_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages

def apply_nl_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages

def apply_kz_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['L3'] = str(fatura_date)
    ws['L4'] = fatura_no
    ws['L5'] = packages

def find_ba_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_ba.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_ba.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_ba.xlsx'),
        os.path.join(os.getcwd(), 'ref_ba.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_ba.xlsx not found. Checked: {candidates}')

def find_ge_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_ge.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_ge.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_ge.xlsx'),
        os.path.join(os.getcwd(), 'ref_ge.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_ge.xlsx not found. Checked: {candidates}')

def find_ko_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_ko.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_ko.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_ko.xlsx'),
        os.path.join(os.getcwd(), 'ref_ko.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_ko.xlsx not found. Checked: {candidates}')

def find_mk_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_mk.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_mk.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_mk.xlsx'),
        os.path.join(os.getcwd(), 'ref_mk.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_mk.xlsx not found. Checked: {candidates}')

def find_be_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_be.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_be.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_be.xlsx'),
        os.path.join(os.getcwd(), 'ref_be.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_be.xlsx not found. Checked: {candidates}')

def find_de_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_de.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_de.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_de.xlsx'),
        os.path.join(os.getcwd(), 'ref_de.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_de.xlsx not found. Checked: {candidates}')

def find_nl_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_nl.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_nl.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_nl.xlsx'),
        os.path.join(os.getcwd(), 'ref_nl.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_nl.xlsx not found. Checked: {candidates}')

def find_kz_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_kz.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_kz.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_kz.xlsx'),
        os.path.join(os.getcwd(), 'ref_kz.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_kz.xlsx not found. Checked: {candidates}')

def find_ru_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_ru.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_ru.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_ru.xlsx'),
        os.path.join(os.getcwd(), 'ref_ru.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_ru.xlsx not found. Checked: {candidates}')

def find_uz_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_uz.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_uz.xlsx'),
        os.path.join(os.path.dirname(current_dir), 'ref_uz.xlsx'),
        os.path.join(os.getcwd(), 'ref_uz.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_uz.xlsx not found. Checked: {candidates}')

def find_iq_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_iq.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_iq.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_iq.xlsx not found.')

def find_ly_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_ly.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_ly.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_ly.xlsx not found.')

def find_lr_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_lr.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_lr.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_lr.xlsx not found.')

def find_lb_template_path():
    current_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(os.path.dirname(current_dir), 'templates', 'ref_lb.xlsx'),
        os.path.join(current_dir, 'templates', 'ref_lb.xlsx'),
    ]
    for path in candidates:
        if os.path.exists(path): return path
    raise FileNotFoundError(f'ref_lb.xlsx not found.')

def apply_ru_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['L3'] = str(fatura_date)
    ws['L4'] = fatura_no
    ws['L5'] = packages

def apply_uz_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['L3'] = str(fatura_date)
    ws['L4'] = fatura_no
    ws['L5'] = packages

def apply_genel_template_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A1'] = sheet_title  # COMMERCIAL INVOICE veya PACKING LIST
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages
# ── EUR tabanlı INV+PL ortak üretim motoru ────────────────────────────────────
def _generate_excel_eur(df, grup_kilolari, hedef_brut, exception_skus,
                        pdf_fields, hedef_net, depo_tipi, eur_kuru,
                        freight_value, insurance_value,
                        find_template_fn, apply_header_fn,
                        inv_cols=None, pl_cols=None, df_original=None):
    if inv_cols is None: inv_cols = KO_INV_COLS
    if pl_cols  is None: pl_cols  = KO_PL_COLS

    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    # Orijinal df için ağırlık hesapla — master Excel için (gruplandırma öncesi)
    df_for_master = df_original if df_original is not None else df
    brut_original, net_original = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    # SKU bazında gruplandırma — INV+PL için
    df = _sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    if not eur_kuru or eur_kuru <= 0:
        eur_kuru = 1.0

    # Gruplandırılmış df için ağırlık hesapla — INV+PL için
    brut_list, net_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)

    if depo_tipi == 'antrepo' and hedef_net > 0:
        toplam_brut = sum(brut_list)
        if toplam_brut > 0:
            net_list_new = []
            toplam_net = 0.0
            for i, b in enumerate(brut_list):
                if i < len(brut_list) - 1:
                    val = round((b / toplam_brut) * hedef_net, 2)
                    net_list_new.append(val)
                    toplam_net += val
                else:
                    net_list_new.append(round(hedef_net - toplam_net, 2))
            net_list = net_list_new

    EUR_FMT       = '#,##0.00 "EUR"'
    INV_TOTAL_COL = 9
    PL_GROSS_COL  = 8
    PL_NET_COL    = 9

    wb = openpyxl.load_workbook(find_template_fn())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']
    DS = 9

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    packages_str = str((pdf_fields or {}).get('kap', '') or '')
    apply_header_fn(ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date, packages_str)
    apply_header_fn(ws_pl,  'PACKING LIST',       fatura_no, fatura_date, packages_str)

    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(inv_cols):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(inv_cols):
            cn = c_idx + 1
            if src_col == '__EUR__':
                birim_eur = parse_num(row.get('Fiyat', 0)) / eur_kuru
                dat(ws_inv, er, cn, birim_eur, bg=bg, align='right', fmt=EUR_FMT)
            elif src_col == '__EUR_CALC__':
                miktar    = parse_num(row.get('Miktar', 0))
                birim_eur = parse_num(row.get('Fiyat', 0)) / eur_kuru
                dat(ws_inv, er, cn, miktar * birim_eur, bg=bg, align='right', fmt=EUR_FMT)
            elif out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE', 'BARCODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_inv = DS + len(df)
    tr, fr, ir, gr = last_inv+1, last_inv+2, last_inv+3, last_inv+4
    for r, h in [(tr, 22), (fr, 22), (ir, 22), (gr, 28)]:
        ws_inv.row_dimensions[r].height = h

    G, H = 8, 9
    tc = get_column_letter(INV_TOTAL_COL)

    c = ws_inv.cell(row=tr, column=G, value='TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=tr, column=H, value=f'=SUM({tc}{DS+1}:{tc}{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = EUR_FMT
    c.border = brd()

    c = ws_inv.cell(row=fr, column=G, value='FREIGHT')
    c.font = Font(name='Arial', bold=True, color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=fr, column=H, value=round(freight_value, 2))
    c.font = Font(name='Arial', color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = EUR_FMT
    c.border = brd()

    c = ws_inv.cell(row=ir, column=G, value='INSURANCE')
    c.font = Font(name='Arial', bold=True, color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=ir, column=H, value=round(insurance_value, 2))
    c.font = Font(name='Arial', color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = EUR_FMT
    c.border = brd()

    c = ws_inv.cell(row=gr, column=G, value='GRAND TOTAL EUR')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=gr, column=H, value=f'=I{tr}+I{fr}+I{ir}')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = EUR_FMT
    c.border = brd()

    set_print(ws_inv, f'A1:L{gr}')

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(pl_cols):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(pl_cols):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 6):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    c = ws_pl.cell(row=pl_gr, column=7, value='TOTAL KG:')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd()

    cl = get_column_letter(PL_GROSS_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_GROSS_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    cl = get_column_letter(PL_NET_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_NET_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    set_print(ws_pl, f'A1:I{pl_gr}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_original, net_original)
    return buf.getvalue(), fatura_no, master_out


def generate_excel_ko(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', eur_kuru=1.0, df_original=None):
    freight_value   = float((pdf_fields or {}).get('navlun',  0) or 0) / eur_kuru
    insurance_value = float((pdf_fields or {}).get('sigorta', 0) or 0) / eur_kuru
    return _generate_excel_eur(
        df, grup_kilolari, hedef_brut, exception_skus,
        pdf_fields, hedef_net, depo_tipi, eur_kuru,
        freight_value, insurance_value,
        find_ko_template_path, apply_ko_template_header,
        df_original=df_original
    )

def generate_excel_mk(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', eur_kuru=1.0, df_original=None):
    freight_value   = float((pdf_fields or {}).get('navlun',  0) or 0) / eur_kuru
    insurance_value = float((pdf_fields or {}).get('sigorta', 0) or 0) / eur_kuru
    return _generate_excel_eur(
        df, grup_kilolari, hedef_brut, exception_skus,
        pdf_fields, hedef_net, depo_tipi, eur_kuru,
        freight_value, insurance_value,
        find_mk_template_path, apply_mk_template_header,
        df_original=df_original
    )

def generate_excel_be(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', eur_kuru=1.0, df_original=None):
    freight_value   = float((pdf_fields or {}).get('navlun',  0) or 0) / eur_kuru
    insurance_value = float((pdf_fields or {}).get('sigorta', 0) or 0) / eur_kuru
    return _generate_excel_eur(
        df, grup_kilolari, hedef_brut, exception_skus,
        pdf_fields, hedef_net, depo_tipi, eur_kuru,
        freight_value, insurance_value,
        find_be_template_path, apply_be_template_header,
        inv_cols=BE_INV_COLS, pl_cols=BE_PL_COLS,
        df_original=df_original
    )

def generate_excel_de(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', eur_kuru=1.0, df_original=None):
    freight_value   = float((pdf_fields or {}).get('navlun',  0) or 0) / eur_kuru
    insurance_value = float((pdf_fields or {}).get('sigorta', 0) or 0) / eur_kuru
    return _generate_excel_eur(
        df, grup_kilolari, hedef_brut, exception_skus,
        pdf_fields, hedef_net, depo_tipi, eur_kuru,
        freight_value, insurance_value,
        find_de_template_path, apply_de_template_header,
        inv_cols=DE_INV_COLS, pl_cols=DE_PL_COLS,
        df_original=df_original
    )

def generate_excel_nl(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', eur_kuru=1.0, df_original=None):
    freight_value   = float((pdf_fields or {}).get('navlun',  0) or 0) / eur_kuru
    insurance_value = float((pdf_fields or {}).get('sigorta', 0) or 0) / eur_kuru
    return _generate_excel_eur(
        df, grup_kilolari, hedef_brut, exception_skus,
        pdf_fields, hedef_net, depo_tipi, eur_kuru,
        freight_value, insurance_value,
        find_nl_template_path, apply_nl_template_header,
        inv_cols=NL_INV_COLS, pl_cols=NL_PL_COLS,
        df_original=df_original
    )


def generate_excel_kz(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Kazakistan INV + PL üretimi."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    # Orijinal df için ağırlık hesapla — master Excel için
    df_for_master = df_original if df_original is not None else df
    brut_original, net_original = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    # SKU bazında gruplandırma — INV+PL için
    df = _sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    # Gruplandırılmış df için ağırlık hesapla
    brut_list, net_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)

    if depo_tipi == 'antrepo' and hedef_net > 0:
        toplam_brut = sum(brut_list)
        if toplam_brut > 0:
            net_list_new = []
            toplam_net = 0.0
            for i, b in enumerate(brut_list):
                if i < len(brut_list) - 1:
                    val = round((b / toplam_brut) * hedef_net, 2)
                    net_list_new.append(val)
                    toplam_net += val
                else:
                    net_list_new.append(round(hedef_net - toplam_net, 2))
            net_list = net_list_new

    freight_value   = float((pdf_fields or {}).get('navlun',  0) or 0)
    insurance_value = float((pdf_fields or {}).get('sigorta', 0) or 0)
    TRY_FMT = '#,##0.00 "TRY"'

    INV_TOTAL_COL = 12
    PL_GROSS_COL  = 11
    PL_NET_COL    = 12

    wb = openpyxl.load_workbook(find_kz_template_path())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']
    DS = 9

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    packages_str = str((pdf_fields or {}).get('kap', '') or '')
    apply_kz_template_header(ws_inv, 'COMMERCIAL INVOICE  / СЧЕТ-ФАКТУРА', fatura_no, fatura_date, packages_str)
    apply_kz_template_header(ws_pl,  'PACKING LIST  / ТОВАРНАЯ НАКЛАДНАЯ', fatura_no, fatura_date, packages_str)

    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(KZ_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(KZ_INV_COLS):
            cn = c_idx + 1
            if src_col == '__CALC__':
                dat(ws_inv, er, cn,
                    round(parse_num(row.get('Miktar', 0)) * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=TRY_FMT)
            elif out_col == 'UNIT':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col == 'UNIT PRICE':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt=TRY_FMT)
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_inv, er, cn, val, bg=bg, align='left')

    last_inv = DS + len(df)
    tr, fr, ir, gr = last_inv+1, last_inv+2, last_inv+3, last_inv+4
    for r, h in [(tr, 22), (fr, 22), (ir, 22), (gr, 28)]:
        ws_inv.row_dimensions[r].height = h

    K, L = 11, 12
    tc = get_column_letter(INV_TOTAL_COL)

    c = ws_inv.cell(row=tr, column=K, value='TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=tr, column=L, value=f'=SUM({tc}{DS+1}:{tc}{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    c = ws_inv.cell(row=fr, column=K, value='FREIGHT')
    c.font = Font(name='Arial', bold=True, color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=fr, column=L, value=freight_value)
    c.font = Font(name='Arial', color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    c = ws_inv.cell(row=ir, column=K, value='INSURANCE')
    c.font = Font(name='Arial', bold=True, color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=ir, column=L, value=insurance_value)
    c.font = Font(name='Arial', color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    c = ws_inv.cell(row=gr, column=K, value='GRAND TOTAL TRY')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=gr, column=L, value=f'=L{tr}+L{fr}+L{ir}')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    set_print(ws_inv, f'A1:P{gr}')

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(KZ_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(KZ_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'UNIT':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_pl, er, cn, val, bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 10):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    c = ws_pl.cell(row=pl_gr, column=10, value='TOTAL KG:')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd()

    cl = get_column_letter(PL_GROSS_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_GROSS_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    cl = get_column_letter(PL_NET_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_NET_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    set_print(ws_pl, f'A1:L{pl_gr}')
    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_original, net_original)
    return buf.getvalue(), fatura_no, master_out

def generate_excel_ru(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Rusya INV + PL üretimi — TRY bazlı, freight/insurance yok."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_original, net_original = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = _sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    brut_list, net_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)

    if depo_tipi == 'antrepo' and hedef_net > 0:
        toplam_brut = sum(brut_list)
        if toplam_brut > 0:
            net_list_new = []
            toplam_net = 0.0
            for i, b in enumerate(brut_list):
                if i < len(brut_list) - 1:
                    val = round((b / toplam_brut) * hedef_net, 2)
                    net_list_new.append(val)
                    toplam_net += val
                else:
                    net_list_new.append(round(hedef_net - toplam_net, 2))
            net_list = net_list_new

    TRY_FMT = '#,##0.00 "TRY"'
    INV_TOTAL_COL = 12  # L
    PL_GROSS_COL  = 11  # K
    PL_NET_COL    = 12  # L

    wb = openpyxl.load_workbook(find_ru_template_path())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']
    DS = 9

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    packages_str = str((pdf_fields or {}).get('kap', '') or '')
    apply_ru_template_header(ws_inv, 'COMMERCIAL INVOICE  / СЧЕТ-ФАКТУРА', fatura_no, fatura_date, packages_str)
    apply_ru_template_header(ws_pl,  'PACKING LIST  / ТОВАРНАЯ НАКЛАДНАЯ', fatura_no, fatura_date, packages_str)

    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(RU_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(RU_INV_COLS):
            cn = c_idx + 1
            if src_col == '__CALC__':
                dat(ws_inv, er, cn,
                    round(parse_num(row.get('Miktar', 0)) * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=TRY_FMT)
            elif out_col == 'UNIT':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col == 'UNIT PRICE':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt=TRY_FMT)
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_inv, er, cn, val, bg=bg, align='left')

    # ── INV footer — sadece TOTAL + GRAND TOTAL (freight/insurance yok) ──────
    last_inv = DS + len(df)
    tr, gr = last_inv + 1, last_inv + 2
    ws_inv.row_dimensions[tr].height = 22
    ws_inv.row_dimensions[gr].height = 28

    K, L = 11, 12
    tc = get_column_letter(INV_TOTAL_COL)

    c = ws_inv.cell(row=tr, column=K, value='TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=tr, column=L, value=f'=SUM({tc}{DS+1}:{tc}{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    c = ws_inv.cell(row=gr, column=K, value='GRAND TOTAL TRY')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=gr, column=L, value=f'=L{tr}')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    set_print(ws_inv, f'A1:P{gr}')

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(RU_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(RU_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'UNIT':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_pl, er, cn, val, bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 10):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    c = ws_pl.cell(row=pl_gr, column=10, value='TOTAL KG:')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd()

    cl = get_column_letter(PL_GROSS_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_GROSS_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    cl = get_column_letter(PL_NET_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_NET_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    set_print(ws_pl, f'A1:L{pl_gr}')
    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_original, net_original)
    return buf.getvalue(), fatura_no, master_out


def generate_excel_uz(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Özbekistan INV + PL üretimi — TRY bazlı, freight/insurance yok."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_original, net_original = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = _sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    brut_list, net_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)

    if depo_tipi == 'antrepo' and hedef_net > 0:
        toplam_brut = sum(brut_list)
        if toplam_brut > 0:
            net_list_new = []
            toplam_net = 0.0
            for i, b in enumerate(brut_list):
                if i < len(brut_list) - 1:
                    val = round((b / toplam_brut) * hedef_net, 2)
                    net_list_new.append(val)
                    toplam_net += val
                else:
                    net_list_new.append(round(hedef_net - toplam_net, 2))
            net_list = net_list_new

    TRY_FMT = '#,##0.00 "TRY"'
    INV_TOTAL_COL = 12  # L
    PL_GROSS_COL  = 11  # K
    PL_NET_COL    = 12  # L

    wb = openpyxl.load_workbook(find_uz_template_path())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']
    DS = 9

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    packages_str = str((pdf_fields or {}).get('kap', '') or '')
    apply_uz_template_header(ws_inv, 'COMMERCIAL INVOICE  / СЧЕТ-ФАКТУРА', fatura_no, fatura_date, packages_str)
    apply_uz_template_header(ws_pl,  'PACKING LIST  / ТОВАРНАЯ НАКЛАДНАЯ', fatura_no, fatura_date, packages_str)

    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(UZ_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(UZ_INV_COLS):
            cn = c_idx + 1
            if src_col == '__CALC__':
                dat(ws_inv, er, cn,
                    round(parse_num(row.get('Miktar', 0)) * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=TRY_FMT)
            elif out_col == 'UNIT':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col == 'UNIT PRICE':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt=TRY_FMT)
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_inv, er, cn, val, bg=bg, align='left')

    # ── INV footer — sadece TOTAL + GRAND TOTAL ───────────────────────────────
    last_inv = DS + len(df)
    tr, gr = last_inv + 1, last_inv + 2
    ws_inv.row_dimensions[tr].height = 22
    ws_inv.row_dimensions[gr].height = 28

    K, L = 11, 12
    tc = get_column_letter(INV_TOTAL_COL)

    c = ws_inv.cell(row=tr, column=K, value='TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=tr, column=L, value=f'=SUM({tc}{DS+1}:{tc}{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    c = ws_inv.cell(row=gr, column=K, value='GRAND TOTAL TRY')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=gr, column=L, value=f'=L{tr}')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    set_print(ws_inv, f'A1:P{gr}')

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(UZ_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(UZ_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'UNIT':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_pl, er, cn, val, bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 10):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    c = ws_pl.cell(row=pl_gr, column=10, value='TOTAL KG:')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd()

    cl = get_column_letter(PL_GROSS_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_GROSS_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    cl = get_column_letter(PL_NET_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_NET_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    set_print(ws_pl, f'A1:L{pl_gr}')
    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_original, net_original)
    return buf.getvalue(), fatura_no, master_out

def _generate_excel_usd(df, grup_kilolari, hedef_brut, exception_skus,
                         pdf_fields, hedef_net, depo_tipi, usd_kuru,
                         find_template_fn, df_original=None):
    """USD bazlı genel INV+PL üretim motoru — Irak, Libya, Liberya, Lübnan."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_original, net_original = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = _sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    if not usd_kuru or usd_kuru <= 0:
        usd_kuru = 1.0

    brut_list, net_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)

    if depo_tipi == 'antrepo' and hedef_net > 0:
        toplam_brut = sum(brut_list)
        if toplam_brut > 0:
            net_list_new = []
            toplam_net = 0.0
            for i, b in enumerate(brut_list):
                if i < len(brut_list) - 1:
                    val = round((b / toplam_brut) * hedef_net, 2)
                    net_list_new.append(val)
                    toplam_net += val
                else:
                    net_list_new.append(round(hedef_net - toplam_net, 2))
            net_list = net_list_new

    USD_FMT       = '#,##0.00 "USD"'
    INV_TOTAL_COL = 9   # I — TOTAL AMOUNT USD
    PL_GROSS_COL  = 8   # H
    PL_NET_COL    = 9   # I

    wb = openpyxl.load_workbook(find_template_fn())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']
    DS = 9  # kolon başlığı satırı — satır 8, veri DS+1'den başlar

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    packages_str = str((pdf_fields or {}).get('kap', '') or '')
    apply_genel_template_header(ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date, packages_str)
    apply_genel_template_header(ws_pl,  'PACKING LIST',       fatura_no, fatura_date, packages_str)

    # INV kolon başlıkları
    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(GENEL_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # INV veri satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(GENEL_INV_COLS):
            cn = c_idx + 1
            if src_col == '__USD__':
                dat(ws_inv, er, cn, parse_num(row.get('Fiyat', 0)), bg=bg, align='right', fmt=USD_FMT)
            elif src_col == '__USD_CALC__':
                miktar = parse_num(row.get('Miktar', 0))
                dat(ws_inv, er, cn, round(miktar * parse_num(row.get('Fiyat', 0)), 2), bg=bg, align='right', fmt=USD_FMT)
            elif out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    # INV footer — GRAND TOTAL USD (freight/insurance yok)
    last_inv = DS + len(df)
    tr, gr = last_inv + 1, last_inv + 2
    ws_inv.row_dimensions[tr].height = 22
    ws_inv.row_dimensions[gr].height = 28

    H, I = 8, 9
    tc = get_column_letter(INV_TOTAL_COL)

    c = ws_inv.cell(row=tr, column=H, value='TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=tr, column=I, value=f'=SUM({tc}{DS+1}:{tc}{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = USD_FMT
    c.border = brd()

    c = ws_inv.cell(row=gr, column=H, value='GRAND TOTAL USD')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=gr, column=I, value=f'=I{tr}')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = USD_FMT
    c.border = brd()

    set_print(ws_inv, f'A1:L{gr}')

    # PL kolon başlıkları
    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(GENEL_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # PL veri satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(GENEL_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    # PL footer — TOTAL KG
    last_pl = DS + len(df)
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 7):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    c = ws_pl.cell(row=pl_gr, column=7, value='TOTAL KG:')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd()

    cl = get_column_letter(PL_GROSS_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_GROSS_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    cl = get_column_letter(PL_NET_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_NET_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    set_print(ws_pl, f'A1:I{pl_gr}')
    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_original, net_original)
    return buf.getvalue(), fatura_no, master_out


def generate_excel_iq(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', usd_kuru=1.0, df_original=None):
    return _generate_excel_usd(df, grup_kilolari, hedef_brut, exception_skus,
                                pdf_fields, hedef_net, depo_tipi, usd_kuru,
                                find_iq_template_path, df_original=df_original)

def generate_excel_ly(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', usd_kuru=1.0, df_original=None):
    return _generate_excel_usd(df, grup_kilolari, hedef_brut, exception_skus,
                                pdf_fields, hedef_net, depo_tipi, usd_kuru,
                                find_ly_template_path, df_original=df_original)

def generate_excel_lr(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', usd_kuru=1.0, df_original=None):
    return _generate_excel_usd(df, grup_kilolari, hedef_brut, exception_skus,
                                pdf_fields, hedef_net, depo_tipi, usd_kuru,
                                find_lr_template_path, df_original=df_original)

def generate_excel_lb(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', usd_kuru=1.0, df_original=None):
    return _generate_excel_usd(df, grup_kilolari, hedef_brut, exception_skus,
                                pdf_fields, hedef_net, depo_tipi, usd_kuru,
                                find_lb_template_path, df_original=df_original)

def generate_excel_ba(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Bosna INV + PL üretimi."""
    df['Birim Cinsi (1)'] = df['Birim Cinsi (1)'].apply(lambda x: 'PCS' if str(x).strip()=='AD' else x)
    df['GTİP'] = df['GTİP'].apply(lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['','nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['','nan'] else '')

    # Orijinal df için ağırlık hesapla — master Excel için
    df_for_master = df_original if df_original is not None else df
    brut_original, net_original = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    # SKU bazında gruplandırma — INV+PL için
    df = _sku_grupla(df)

    fatura_no    = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date  = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'): fatura_date = fatura_date.date()
    musteri      = 'Madame Coco BH d.o.o.'
    musteri_adres= 'Ulica Vrbanja Br. 1 (SCC)-Saraybosna, Saraybosna Centar'
    destination  = 'Bosnia and Herzegovina'
    incoterm     = 'CIP'

    # Gruplandırılmış df için ağırlık hesapla
    brut_list, net_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)

    if depo_tipi == 'antrepo' and hedef_net > 0:
        toplam_brut = sum(brut_list)
        if toplam_brut > 0:
            net_list_new = []
            toplam_net = 0.0
            for i, b in enumerate(brut_list):
                if i < len(brut_list) - 1:
                    val = round((b / toplam_brut) * hedef_net, 2)
                    net_list_new.append(val)
                    toplam_net += val
                else:
                    net_list_new.append(round(hedef_net - toplam_net, 2))
            net_list = net_list_new

    INV_UNIT_COL  = 7
    INV_TOTAL_COL = 8
    PL_GROSS_COL  = 7
    PL_NET_COL    = 8

    wb = openpyxl.load_workbook(find_ba_template_path())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']
    DS = 9

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    apply_ba_template_header(
        ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date,
        musteri, musteri_adres, destination, incoterm,
        packages=str((pdf_fields or {}).get('kap', ''))
    )
    apply_ba_template_header(
        ws_pl, 'PACKING LIST', fatura_no, fatura_date,
        musteri, musteri_adres, destination, incoterm,
        packages=str((pdf_fields or {}).get('kap', ''))
    )

    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(BA_INV_COLS):
        hdr(ws_inv, DS, i+1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(BA_INV_COLS):
            cn = c_idx + 1
            if out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('UNIT PRICE', 'TOTAL AMOUNT TRY'):
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0.00 "TRY"')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_inv = DS + len(df)
    gr = last_inv + 1
    ws_inv.row_dimensions[gr].height = 28

    c = ws_inv.cell(row=gr, column=INV_UNIT_COL, value='GRAND TOTAL TRY')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()

    tc = get_column_letter(INV_TOTAL_COL)
    c = ws_inv.cell(row=gr, column=INV_TOTAL_COL,
                    value=f'=SUM({tc}{DS+1}:{tc}{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00 "TRY"'
    c.border = brd()

    set_print(ws_inv, f'A1:H{gr}')

    for col, w in [('A',16),('B',14),('C',13),('D',18.18),('E',33),('F',6),
                   ('G',15.54),('H',20.91)]:
        ws_pl.column_dimensions[col].width = w

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(BA_PL_COLS):
        hdr(ws_pl, DS, i+1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(BA_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col == 'MASTER ITEM CODE':
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 6):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    c = ws_pl.cell(row=pl_gr, column=6, value='TOTAL KG:')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd()

    cl = get_column_letter(PL_GROSS_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_GROSS_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    cl = get_column_letter(PL_NET_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_NET_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    set_print(ws_pl, f'A1:H{pl_gr}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_original, net_original)
    return buf.getvalue(), fatura_no, master_out


def generate_excel_ge(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                      pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Gürcistan INV + PL üretimi."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    # Orijinal df için ağırlık hesapla — master Excel için
    df_for_master = df_original if df_original is not None else df
    brut_original, net_original = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    # SKU bazında gruplandırma — INV+PL için
    df = _sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    # Gruplandırılmış df için ağırlık hesapla
    brut_list, net_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)

    if depo_tipi == 'antrepo' and hedef_net > 0:
        toplam_brut = sum(brut_list)
        if toplam_brut > 0:
            net_list_new = []
            toplam_net = 0.0
            for i, b in enumerate(brut_list):
                if i < len(brut_list) - 1:
                    val = round((b / toplam_brut) * hedef_net, 2)
                    net_list_new.append(val)
                    toplam_net += val
                else:
                    net_list_new.append(round(hedef_net - toplam_net, 2))
            net_list = net_list_new

    INV_TOTAL_COL = 8
    PL_GROSS_COL  = 7
    PL_NET_COL    = 8
    TRY_FMT       = '#,##0.00 "TRY"'

    wb = openpyxl.load_workbook(find_ge_template_path())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']
    DS = 9

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    packages_str    = str((pdf_fields or {}).get('kap', '') or '')
    freight_value   = parse_num((pdf_fields or {}).get('navlun', 0))
    insurance_value = parse_num((pdf_fields or {}).get('sigorta', 0))

    apply_ge_template_header(ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date, packages_str)
    apply_ge_template_header(ws_pl,  'PACKING LIST',       fatura_no, fatura_date, packages_str)
    ws_inv['H5'] = packages_str
    ws_pl['H5']  = packages_str

    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(GE_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(GE_INV_COLS):
            cn = c_idx + 1
            if out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif src_col == '__CALC__':
                dat(ws_inv, er, cn,
                    round(parse_num(row.get('Miktar', 0)) * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=TRY_FMT)
            elif out_col == 'UNIT PRICE':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt=TRY_FMT)
            elif out_col in ('MASTER ITEM CODE', 'HS CODE', 'BARCODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_inv = DS + len(df)
    tr, fr, ir, gr = last_inv+1, last_inv+2, last_inv+3, last_inv+4
    for r, h in [(tr, 22), (fr, 22), (ir, 22), (gr, 28)]:
        ws_inv.row_dimensions[r].height = h

    G, H = 7, 8
    tc = get_column_letter(INV_TOTAL_COL)

    c = ws_inv.cell(row=tr, column=G, value='TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=tr, column=H, value=f'=SUM({tc}{DS+1}:{tc}{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    c = ws_inv.cell(row=fr, column=G, value='FREIGHT')
    c.font = Font(name='Arial', bold=True, color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=fr, column=H, value=freight_value)
    c.font = Font(name='Arial', color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    c = ws_inv.cell(row=ir, column=G, value='INSURANCE')
    c.font = Font(name='Arial', bold=True, color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=ir, column=H, value=insurance_value)
    c.font = Font(name='Arial', color='000000', size=9)
    c.fill = PatternFill('solid', fgColor='FFFFFF')
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    c = ws_inv.cell(row=gr, column=G, value='GRAND TOTAL TRY')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=gr, column=H, value=f'=H{tr}+H{fr}+H{ir}')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TRY_FMT
    c.border = brd()

    set_print(ws_inv, f'A1:L{gr}')

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(GE_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(GE_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 6):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    c = ws_pl.cell(row=pl_gr, column=6, value='TOTAL KG:')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd()

    cl = get_column_letter(PL_GROSS_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_GROSS_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    cl = get_column_letter(PL_NET_COL)
    c = ws_pl.cell(row=pl_gr, column=PL_NET_COL,
                   value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00'
    c.border = brd()

    set_print(ws_pl, f'A1:H{pl_gr}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_original, net_original)
    return buf.getvalue(), fatura_no, master_out


def generate_excel(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                   pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Sırbistan INV + PL üretimi."""
    df['Birim Cinsi (1)'] = df['Birim Cinsi (1)'].apply(
        lambda x: 'PCS' if str(x).strip()=='AD' else x)
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['','nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['','nan'] else '')

    # Orijinal df için ağırlık hesapla — master Excel için
    df_for_master = df_original if df_original is not None else df
    brut_original, net_original = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    # SKU bazında gruplandırma — INV+PL için
    df = _sku_grupla(df)

    fatura_no    = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date  = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date,'date'): fatura_date = fatura_date.date()
    musteri      = str(df['Müşteri Firma Adı'].iloc[0]).strip()
    musteri_adres= 'Gospodar Jovanova 73, Belgrade'

    if pdf_fields is None: pdf_fields = {'navlun': 0.0, 'sigorta': 0.0, 'kap': ''}

    # Gruplandırılmış df için ağırlık hesapla
    brut_list, net_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)

    if depo_tipi == 'antrepo' and hedef_net > 0:
        toplam_brut = sum(brut_list)
        if toplam_brut > 0:
            net_list = []
            toplam_yuvarlanmis_net = 0.0
            for i, b in enumerate(brut_list):
                if i < len(brut_list) - 1:
                    val = round((b / toplam_brut) * hedef_net, 2)
                    net_list.append(val)
                    toplam_yuvarlanmis_net += val
                else:
                    net_list.append(round(hedef_net - toplam_yuvarlanmis_net, 2))

    wb = Workbook()
    DS = 9

    ws_inv = wb.active
    ws_inv.title = 'INV'
    build_header(ws_inv,'COMMERCIAL INVOICE',fatura_no,fatura_date,musteri,musteri_adres,len(INV_COLS),logo_bytes,pdf_fields)
    ws_inv.row_dimensions[DS].height = 35
    for i,(hd,_) in enumerate(INV_COLS):
        hdr(ws_inv,DS,i+1,hd,bg=DARK_BLUE,size=9,align='center')
    for r_idx,(_,row) in enumerate(df.iterrows()):
        er=DS+1+r_idx; ws_inv.row_dimensions[er].height=None
        bg='FFFFFF' if r_idx%2==0 else 'EBF3FB'
        for c_idx,(out_col,src_col) in enumerate(INV_COLS):
            cn=c_idx+1
            if src_col=='__CALC__':
                dat(ws_inv,er,cn,round(parse_num(row.get('Miktar',0))*parse_num(row.get('Fiyat',0)),2),bg=bg,align='right',fmt=TL_FMT)
            elif out_col=='QTY':
                dat(ws_inv,er,cn,parse_num(row.get(src_col,0)),bg=bg,align='right',fmt='#,##0')
            elif out_col=='UNIT PRICE':
                dat(ws_inv,er,cn,parse_num(row.get(src_col,0)),bg=bg,align='right',fmt=TL_FMT)
            elif out_col in('MASTER ITEM CODE','HS CODE'):
                dat(ws_inv,er,cn,str(row.get(src_col,'') or ''),bg=bg,align='left')
            else:
                dat(ws_inv,er,cn,row.get(src_col,''),bg=bg,align='left')

    last_inv=DS+len(df)
    tr,fr,ir,gr=last_inv+1,last_inv+2,last_inv+3,last_inv+4
    for r in[tr,fr,ir,gr]: ws_inv.row_dimensions[r].height=22
    ws_inv.row_dimensions[gr].height=28
    H,I=8,9
    ws_inv.merge_cells(f'A{tr}:G{tr}')
    c=ws_inv.cell(row=tr,column=H,value='TOTAL')
    c.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
    c.fill=PatternFill('solid',fgColor=DARK_BLUE)
    c.alignment=Alignment(horizontal='center',vertical='center'); c.border=brd()
    c=ws_inv.cell(row=tr,column=I,value=f'=SUM(I{DS+1}:I{last_inv})')
    c.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
    c.fill=PatternFill('solid',fgColor=DARK_BLUE)
    c.alignment=Alignment(horizontal='right',vertical='center')
    c.number_format=TL_FMT; c.border=brd()
    ws_inv.merge_cells(f'A{fr}:G{fr}')
    dat(ws_inv,fr,H,'FREIGHT',bold=True,align='center')
    dat(ws_inv,fr,I,float(pdf_fields.get('navlun',0)),fmt=TL_FMT,align='right')
    ws_inv.merge_cells(f'A{ir}:G{ir}')
    dat(ws_inv,ir,H,'INSURANCE',bold=True,align='center')
    dat(ws_inv,ir,I,float(pdf_fields.get('sigorta',0)),fmt=TL_FMT,align='right')
    ws_inv.merge_cells(f'A{gr}:G{gr}')
    c=ws_inv.cell(row=gr,column=H,value='GRAND TOTAL')
    c.font=Font(name='Arial',bold=True,color='FFFFFF',size=11)
    c.fill=PatternFill('solid',fgColor=GOLD)
    c.alignment=Alignment(horizontal='center',vertical='center'); c.border=brd()
    c=ws_inv.cell(row=gr,column=I,value=f'=I{tr}+I{fr}+I{ir}')
    c.font=Font(name='Arial',bold=True,color='FFFFFF',size=11)
    c.fill=PatternFill('solid',fgColor=GOLD)
    c.alignment=Alignment(horizontal='right',vertical='center')
    c.number_format=TL_FMT; c.border=brd()
    build_footer(ws_inv,gr+1,len(INV_COLS))
    set_print(ws_inv,f'A1:I{gr}')

    ws_pl=wb.create_sheet('PL')
    pl_widths={'A':16,'B':14,'C':14,'D':18,'E':33,'F':6,'G':7,'H':14,'I':14}
    for col,w in pl_widths.items(): ws_pl.column_dimensions[col].width=w
    build_header(ws_pl,'PACKING LIST',fatura_no,fatura_date,musteri,musteri_adres,len(PL_COLS),logo_bytes,pdf_fields)
    ws_pl.row_dimensions[DS].height=35
    for i,(hd,_) in enumerate(PL_COLS):
        hdr(ws_pl,DS,i+1,hd,bg=DARK_BLUE,size=9,align='center')
    for r_idx,(_,row) in enumerate(df.iterrows()):
        er=DS+1+r_idx; ws_pl.row_dimensions[er].height=None
        bg='FFFFFF' if r_idx%2==0 else 'EBF3FB'
        for c_idx,(out_col,src_col) in enumerate(PL_COLS):
            cn=c_idx+1
            if src_col=='__BRUT__':
                dat(ws_pl,er,cn,round(brut_list[r_idx],2),bg=bg,align='right',fmt='#,##0.00')
            elif src_col=='__NET__':
                dat(ws_pl,er,cn,round(net_list[r_idx],2),bg=bg,align='right',fmt='#,##0.00')
            elif out_col=='QTY':
                dat(ws_pl,er,cn,parse_num(row.get(src_col,0)),bg=bg,align='right',fmt='#,##0')
            elif out_col=='MASTER ITEM CODE':
                dat(ws_pl,er,cn,str(row.get(src_col,'') or ''),bg=bg,align='left')
            else:
                dat(ws_pl,er,cn,row.get(src_col,''),bg=bg,align='left')

    last_pl=DS+len(df)
    pl_gr=last_pl+1
    ws_pl.row_dimensions[pl_gr].height=28
    for col_idx in range(1, 5):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')
    ws_pl.merge_cells(f'E{pl_gr}:F{pl_gr}')
    c=ws_pl.cell(row=pl_gr,column=5,value='GRAND TOTAL')
    c.font=Font(name='Arial',bold=True,color='FFFFFF',size=11)
    c.fill=PatternFill('solid',fgColor=GOLD)
    c.alignment=Alignment(horizontal='center',vertical='center'); c.border=brd()
    for cn,fmt in[(7,'#,##0'),(8,'#,##0.00'),(9,'#,##0.00')]:
        cl=get_column_letter(cn)
        c=ws_pl.cell(row=pl_gr,column=cn,value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
        c.font=Font(name='Arial',bold=True,color='FFFFFF',size=11)
        c.fill=PatternFill('solid',fgColor=GOLD)
        c.alignment=Alignment(horizontal='right',vertical='center')
        c.number_format=fmt; c.border=brd()
    build_footer(ws_pl,pl_gr+1,len(PL_COLS))
    set_print(ws_pl,f'A1:I{pl_gr}')

    from openpyxl.worksheet.properties import WorksheetProperties
    for sheet in wb.worksheets:
        sheet.sheet_properties.enableFormatConditionsCalculation = True
    wb.calculation.calcMode = 'auto'

    buf=io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_original, net_original)
    return buf.getvalue(), fatura_no, master_out


# ── VERCEL HANDLER ────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Access-Control-Allow-Headers','Content-Type')
        self.send_header('Access-Control-Allow-Methods','POST, OPTIONS')
        self.end_headers()

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = json.loads(self.rfile.read(length))

            excel_bytes    = base64.b64decode(body.get('excel',''))
            logo_b64       = body.get('logo','')
            logo_bytes     = base64.b64decode(logo_b64) if logo_b64 else None
            hedef_brut     = float(body.get('hedefBrut', 0))
            hedef_net      = float(body.get('hedefNet', 0))
            depo_tipi      = body.get('depoTipi', 'serbest')
            grup_kilolari  = body.get('grupKilolari', {})
            exception_skus = body.get('exceptionSkus', EXCEPTION_SKUS)

            pdf_b64    = body.get('pdf', '')
            pdf_fields = {'navlun': 0.0, 'sigorta': 0.0, 'kap': ''}
            if pdf_b64:
                pdf_bytes_data = base64.b64decode(pdf_b64)
                pdf_fields = parse_pdf(pdf_bytes_data)

            ulke_kodu   = body.get('ulkeKodu', 'rs')
            df          = pd.read_excel(io.BytesIO(excel_bytes))
            df_original = df.copy()

            if ulke_kodu == 'ba':
                excel_out, fatura_no, master_out = generate_excel_ba(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, df_original=df_original)
            elif ulke_kodu == 'ge':
                excel_out, fatura_no, master_out = generate_excel_ge(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, df_original=df_original)
            elif ulke_kodu == 'xk':
                eur_kuru = float(body.get('eurKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_ko(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, eur_kuru=eur_kuru,
                    df_original=df_original)
            elif ulke_kodu == 'mk':
                eur_kuru = float(body.get('eurKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_mk(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, eur_kuru=eur_kuru,
                    df_original=df_original)
            elif ulke_kodu == 'kz':
                excel_out, fatura_no, master_out = generate_excel_kz(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, df_original=df_original)
            elif ulke_kodu == 'ru':
                excel_out, fatura_no, master_out = generate_excel_ru(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, df_original=df_original)
            elif ulke_kodu == 'uz':
                excel_out, fatura_no, master_out = generate_excel_uz(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, df_original=df_original)    
            elif ulke_kodu == 'be':
                eur_kuru = float(body.get('eurKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_be(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, eur_kuru=eur_kuru,
                    df_original=df_original)
            elif ulke_kodu == 'de':
                eur_kuru = float(body.get('eurKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_de(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, eur_kuru=eur_kuru,
                    df_original=df_original)
            elif ulke_kodu == 'nl':
                eur_kuru = float(body.get('eurKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_nl(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, eur_kuru=eur_kuru,
                    df_original=df_original)
            elif ulke_kodu == 'iq':
                usd_kuru = float(body.get('usdKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_iq(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, usd_kuru=usd_kuru,
                    df_original=df_original)
            elif ulke_kodu == 'ly':
                usd_kuru = float(body.get('usdKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_ly(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, usd_kuru=usd_kuru,
                    df_original=df_original)
            elif ulke_kodu == 'lr':
                usd_kuru = float(body.get('usdKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_lr(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, usd_kuru=usd_kuru,
                    df_original=df_original)
            elif ulke_kodu == 'lb':
                usd_kuru = float(body.get('usdKuru', 1.0))
                excel_out, fatura_no, master_out = generate_excel_lb(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, usd_kuru=usd_kuru,
                    df_original=df_original)
            else:
                excel_out, fatura_no, master_out = generate_excel(
                    df, grup_kilolari, hedef_brut, exception_skus, logo_bytes, pdf_fields,
                    hedef_net=hedef_net, depo_tipi=depo_tipi, df_original=df_original)

            result = json.dumps({
                'success':   True,
                'excel':     base64.b64encode(excel_out).decode('utf-8'),
                'master':    base64.b64encode(master_out).decode('utf-8'),
                'faturaNo':  fatura_no,
                'pdfFields': pdf_fields,
            })
            self.send_response(200)
            self.send_header('Content-Type','application/json')
            self.send_header('Access-Control-Allow-Origin','*')
            self.end_headers()
            self.wfile.write(result.encode('utf-8'))

        except Exception as e:
            err = json.dumps({'success':False,'error':str(e),'trace':traceback.format_exc()})
            self.send_response(500)
            self.send_header('Content-Type','application/json')
            self.send_header('Access-Control-Allow-Origin','*')
            self.end_headers()
            self.wfile.write(err.encode('utf-8'))