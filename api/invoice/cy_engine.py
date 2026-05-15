"""
Kıbrıs özel engine — INV YOK, sadece PL üretilir.
1-3 fatura alır, tek PL çıktısı üretir.
"""
import base64
import io

import openpyxl
import pandas as pd
import pdfplumber

from .constants import DARK_BLUE, GOLD, CY_PL_COLS
from .helpers   import hdr, dat, parse_num, sku_grupla, set_print, brd
from .weights   import calculate_weights, get_net_list
from .templates import find_cy_template_path, apply_cy_header

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils  import get_column_letter

DS = 9   # Data start satırı


def _parse_cy_pdf(pdf_bytes):
    """Kıbrıs PDF'inden BRÜT/NET kg ve kap bilgisini çıkarır."""
    import re
    result = {'brutKg': 0.0, 'netKg': 0.0, 'kap': ''}
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = ' '.join(
                (p.extract_text() or '') for p in pdf.pages[-2:]
            )

        def _ext(t, pats):
            for p in pats:
                m = re.search(p, t, re.IGNORECASE)
                if m:
                    try:
                        return float(m.group(1).replace(',', '.'))
                    except Exception:
                        pass
            return 0.0

        result['brutKg'] = _ext(text, [r'\bB\.KG\s*[:.]?\s*([\d.,]+)'])
        result['netKg']  = _ext(text, [r'\bN\.KG\s*[:.]?\s*([\d.,]+)'])

        # Kap sayısı
        from .helpers import _extract_pdf_packages
        result['kap'] = _extract_pdf_packages(text)
    except Exception:
        pass
    return result


def generate_cy(faturalar, grup_kilolari, exception_skus):
    """
    Kıbrıs PL üretimi.

    faturalar: list of dict
        {
          'excel': '<base64>',
          'pdf':   '<base64>' | None,
          'faturaNo': str      # opsiyonel — df'ten okunur
        }
    """
    fatura_list = []

    for f in faturalar:
        excel_bytes = base64.b64decode(f['excel'])
        df = pd.read_excel(io.BytesIO(excel_bytes), engine='openpyxl')

        # PDF'ten ağırlık ve kap bilgisi
        pdf_fields = {'brutKg': 0.0, 'netKg': 0.0, 'kap': ''}
        if f.get('pdf'):
            pdf_bytes  = base64.b64decode(f['pdf'])
            pdf_fields = _parse_cy_pdf(pdf_bytes)

        hedef_brut = float(pdf_fields.get('brutKg', 0) or 0)
        hedef_net  = float(pdf_fields.get('netKg',  0) or 0)

        fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
        fatura_date = df['Fatura Tarihi'].iloc[0]
        if hasattr(fatura_date, 'date'):
            fatura_date = fatura_date.date()

        # SKU gruplandır
        df = sku_grupla(df)

        # Ağırlık hesapla
        brut_list, _ = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)
        net_list     = get_net_list(brut_list, hedef_net,
                                    'antrepo' if hedef_net > 0 else 'serbest')

        fatura_list.append({
            'df':          df,
            'brut_list':   brut_list,
            'net_list':    net_list,
            'fatura_no':   fatura_no,
            'fatura_date': fatura_date,
            'kap':         pdf_fields.get('kap', ''),
        })

    # Fatura no'ya göre sırala
    fatura_list.sort(key=lambda x: x['fatura_no'])

    # Şablonu yükle
    wb = openpyxl.load_workbook(find_cy_template_path())
    ws = wb['PL']

    if ws.max_row > DS:
        ws.delete_rows(DS + 1, ws.max_row - DS)

    # Header
    fatura_nos  = ' / '.join(f['fatura_no'] for f in fatura_list)
    toplam_kap  = ' / '.join(str(f['kap']) for f in fatura_list if f['kap'])
    apply_cy_header(ws, fatura_nos, fatura_list[0]['fatura_date'], toplam_kap)

    # Kolon başlıkları
    ws.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(CY_PL_COLS):
        hdr(ws, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # Her faturanın satırlarını alt alta yaz
    current_row = DS + 1

    for f in fatura_list:
        df_f      = f['df']
        brut_list = f['brut_list']
        net_list  = f['net_list']

        for r_idx, (_, row) in enumerate(df_f.iterrows()):
            ws.row_dimensions[current_row].height = 23
            bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'

            for c_idx, (out_col, src_col) in enumerate(CY_PL_COLS):
                cn = c_idx + 1
                if src_col == '__BRUT__':
                    dat(ws, current_row, cn, round(brut_list[r_idx], 2),
                        bg=bg, align='right', fmt='#,##0.00')
                elif src_col == '__NET__':
                    dat(ws, current_row, cn, round(net_list[r_idx], 2),
                        bg=bg, align='right', fmt='#,##0.00')
                elif out_col == 'TOPLAM ÜRÜN ADEDİ':
                    dat(ws, current_row, cn, parse_num(row.get(src_col, 0)),
                        bg=bg, align='right', fmt='#,##0')
                elif out_col == 'Asorti Barkodu':
                    dat(ws, current_row, cn, str(row.get(src_col, '') or ''),
                        bg=bg, align='left')
                else:
                    dat(ws, current_row, cn, row.get(src_col, ''), bg=bg, align='left')

            current_row += 1

    # TOTAL KG footer
    last_row  = current_row - 1
    total_row = current_row
    ws.row_dimensions[total_row].height = 28

    for col_idx in range(1, 6):
        ws.cell(row=total_row, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    def _gold(ws, r, col, val, fmt=None):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        c.fill      = PatternFill('solid', fgColor=GOLD)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border    = brd()
        if fmt:
            c.number_format = fmt
        return c

    _gold(ws, total_row, 6, 'TOTAL KG:')
    _gold(ws, total_row, 7, f'=SUM(G{DS+1}:G{last_row})', fmt='#,##0.00')
    _gold(ws, total_row, 8, f'=SUM(H{DS+1}:H{last_row})', fmt='#,##0.00')

    ws.sheet_view.topLeftCell = 'A1'
    set_print(ws, f'A1:J{total_row}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
