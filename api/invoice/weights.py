import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .constants import DARK_BLUE
from .helpers  import parse_num


def calculate_weights(df, grup_kilolari, hedef_brut, exception_skus):
    """
    Her satır için brüt ve net kg hesaplar.
    - SKU istisna listesindeyse o ağırlığı kullan
    - Ürün ağırlığı varsa onu kullan
    - Yoksa grup ağırlığını kullan
    Toplam hedef_brut'a göre orantılanır, son satır yuvarlama farkını alır.
    Net = brüt × 0.9
    """
    ham_list = []
    for _, row in df.iterrows():
        sku    = str(row.get('SKU', '')).strip()
        grup   = str(row.get('ÜRÜN ARA GRUBU', '')).strip()
        ag     = parse_num(row.get('Ürün Ağırlığı (KG)', 0))
        miktar = parse_num(row.get('Miktar', 0))

        if sku in exception_skus:
            kg = parse_num(exception_skus[sku])
        elif ag > 0:
            kg = ag
        else:
            kg = parse_num(grup_kilolari.get(grup, 0))

        ham_list.append(kg * miktar)

    ham_toplam = sum(ham_list)
    if ham_toplam <= 0:
        return [0.0] * len(ham_list), [0.0] * len(ham_list)

    carpan = hedef_brut / ham_toplam

    # Brüt — son satır yuvarlama farkını alır
    brut_list = []
    toplam_yuvarlanmis = 0.0
    for i, h in enumerate(ham_list):
        if i < len(ham_list) - 1:
            val = round(h * carpan, 2)
            brut_list.append(val)
            toplam_yuvarlanmis += val
        else:
            brut_list.append(round(hedef_brut - toplam_yuvarlanmis, 2))

    # Net = brüt × 0.9, aynı yuvarlama mantığı
    hedef_net_serbest = round(hedef_brut * 0.9, 2)
    net_list  = []
    toplam_net = 0.0
    for i, b in enumerate(brut_list):
        if i < len(brut_list) - 1:
            val = round(b * 0.9, 2)
            net_list.append(val)
            toplam_net += val
        else:
            net_list.append(round(hedef_net_serbest - toplam_net, 2))

    return brut_list, net_list


def _dagit_net(brut_list, hedef_net_toplam):
    """
    Verilen hedef_net_toplam'ı brüt oranına göre satırlara dağıtır.
    Son satır yuvarlama farkını alır — toplam her zaman hedef_net_toplam'a eşit.
    """
    toplam_brut = sum(brut_list)
    if toplam_brut <= 0:
        return [0.0] * len(brut_list)

    net_list   = []
    toplam_net = 0.0
    for i, b in enumerate(brut_list):
        if i < len(brut_list) - 1:
            val = round((b / toplam_brut) * hedef_net_toplam, 2)
            net_list.append(val)
            toplam_net += val
        else:
            net_list.append(round(hedef_net_toplam - toplam_net, 2))
    return net_list


def get_net_list(brut_list, hedef_net, depo_tipi, hedef_brut=None):
    """
    Depo tipine göre net listesini döner.

    Serbest depo:
      hedef_net_toplam = round(hedef_brut * 0.9, 2)  ← sabit hedef
      Bu hedef brüt oranına göre satırlara dağıtılır.
      Böylece PL ve Master her zaman aynı toplamı gösterir.

    Antrepo:
      PDF'ten gelen hedef_net brüt oranına göre dağıtılır.
    """
    if depo_tipi == 'antrepo' and hedef_net > 0:
        return _dagit_net(brut_list, hedef_net)

    # Serbest depo — hedef_brut verilmişse ondan hesapla,
    # verilmemişse brüt toplamından türet (geriye dönük uyumluluk)
    if hedef_brut is not None and hedef_brut > 0:
        hedef_net_toplam = round(hedef_brut * 0.9, 2)
    else:
        hedef_net_toplam = round(sum(brut_list) * 0.9, 2)

    return _dagit_net(brut_list, hedef_net_toplam)


def generate_master_excel(df_original, brut_list, net_list,
                           hedef_net=0, depo_tipi='serbest'):
    """
    Stilize master Excel — header koyu mavi, zebra satırlar,
    freeze panes, auto filter, auto kolon genişliği.

    net_list dışarıdan get_net_list() ile hesaplanmış olarak gelir.
    Böylece PL ve Master her zaman aynı net toplamını gösterir.
    """
    df = df_original.copy()
    df['BRÜT'] = brut_list
    df['NET']  = net_list   # PL ile aynı liste — tutarlı toplam garantili

    # Ürün Ağırlığı (KG) = BRÜT / Miktar
    miktar_arr = df['Miktar'].apply(parse_num)
    brut_arr   = df['BRÜT'].apply(parse_num)
    df['Ürün Ağırlığı (KG)'] = (
        brut_arr / miktar_arr.replace(0, float('nan'))
    ).round(6).fillna(0)

    # BRÜT ve NET kolonlarını Ürün Ağırlığı (KG)'nın hemen arkasına taşı
    cols   = list(df.columns)
    ag_idx = cols.index('Ürün Ağırlığı (KG)')
    for col in ('BRÜT', 'NET'):
        cols.remove(col)
    cols.insert(ag_idx + 1, 'BRÜT')
    cols.insert(ag_idx + 2, 'NET')
    df = df[cols]

    headers   = list(df.columns)
    data_rows = df.values.tolist()
    n_cols    = len(headers)
    ZEBRA_CLR = 'F7FAFD'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Master'

    # Auto kolon genişliği (ilk 200 satır örneklenir)
    sample = min(200, len(data_rows))
    for c_idx in range(n_cols):
        max_len = len(str(headers[c_idx] or ''))
        for row in data_rows[:sample]:
            v = row[c_idx]
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(c_idx + 1)].width = max(10, min(40, max_len + 2))

    # Header satırı
    thin       = Side(style='thin', color='BFBFBF')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font   = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    hdr_fill   = PatternFill('solid', fgColor=DARK_BLUE)
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for c_idx, h in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=c_idx, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border_all
    ws.row_dimensions[1].height = 30

    # Sayısal kolonları tespit et (sağa hizalanacak)
    right_cols = set()
    if data_rows:
        for c_idx, v in enumerate(data_rows[0]):
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                right_cols.add(c_idx)

    # Data satırları
    data_font   = Font(name='Arial', size=9)
    align_left  = Alignment(horizontal='left',  vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    fill_white  = PatternFill('solid', fgColor='FFFFFF')
    fill_zebra  = PatternFill('solid', fgColor=ZEBRA_CLR)

    for r_idx, row in enumerate(data_rows):
        excel_row = r_idx + 2
        bg = fill_white if r_idx % 2 == 0 else fill_zebra
        for c_idx, val in enumerate(row, start=1):
            cell           = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font      = data_font
            cell.fill      = bg
            cell.alignment = align_right if (c_idx - 1) in right_cols else align_left
            cell.border    = border_all

    # Freeze + auto filter
    ws.freeze_panes   = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(n_cols)}{len(data_rows) + 1}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
