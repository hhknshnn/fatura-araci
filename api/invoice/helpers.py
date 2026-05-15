import io
import re

import pdfplumber
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from .constants import DARK_BLUE, MID_BLUE, LIGHT_BLUE, LIGHT_GRAY

# ── Stil cache'leri (tekrar tekrar obje üretmemek için) ───────────────────────
_FONT_CACHE   = {}
_FILL_CACHE   = {}
_BORDER_CACHE = {}
_ALIGN_CACHE  = {}


def brd(c='BFBFBF'):
    """İnce kenarlık — tüm kenarlar."""
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)


def hdr(ws, r, col, val,
        bg=DARK_BLUE, fg='FFFFFF', bold=True, align='center', size=9):
    """Header hücresi yaz."""
    cell = ws.cell(row=r, column=col, value=val)

    fkey = (bold, fg, size)
    if fkey not in _FONT_CACHE:
        _FONT_CACHE[fkey] = Font(name='Arial', bold=bold, color=fg, size=size)
    cell.font = _FONT_CACHE[fkey]

    if bg not in _FILL_CACHE:
        _FILL_CACHE[bg] = PatternFill('solid', fgColor=bg)
    cell.fill = _FILL_CACHE[bg]

    if align not in _ALIGN_CACHE:
        _ALIGN_CACHE[align] = Alignment(
            horizontal=align, vertical='center', wrap_text=True)
    cell.alignment = _ALIGN_CACHE[align]

    if 'default' not in _BORDER_CACHE:
        _BORDER_CACHE['default'] = brd()
    cell.border = _BORDER_CACHE['default']

    return cell


def dat(ws, r, col, val,
        bg='FFFFFF', bold=False, align='left', fmt=None):
    """Data hücresi yaz."""
    cell = ws.cell(row=r, column=col, value=val)

    fkey = (bold,)
    if fkey not in _FONT_CACHE:
        _FONT_CACHE[fkey] = Font(name='Arial', bold=bold, color='000000', size=9)
    cell.font = _FONT_CACHE[fkey]

    if bg not in _FILL_CACHE:
        _FILL_CACHE[bg] = PatternFill('solid', fgColor=bg)
    cell.fill = _FILL_CACHE[bg]

    if align not in _ALIGN_CACHE:
        _ALIGN_CACHE[align] = Alignment(
            horizontal=align, vertical='center', wrap_text=True)
    cell.alignment = _ALIGN_CACHE[align]

    if 'default' not in _BORDER_CACHE:
        _BORDER_CACHE['default'] = brd()
    cell.border = _BORDER_CACHE['default']

    if fmt:
        cell.number_format = fmt

    return cell


def parse_num(v):
    """Herhangi bir değeri float'a çevir. Hata durumunda 0.0 döner."""
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v) if str(v) not in ['nan', 'inf'] else 0.0
    s = str(v).strip().replace(' ', '').replace('\u00a0', '')
    if '.' in s and ',' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except Exception:
        return 0.0


# ── PDF parse yardımcıları ────────────────────────────────────────────────────

def _normalize_pdf_text(text):
    return re.sub(r'\s+', ' ', (text or '').replace('\u00a0', ' ')).strip()


def _parse_pdf_amount(value):
    s = str(value).strip().replace(' ', '').replace('\u00a0', '')
    if '.' in s and ',' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except Exception:
        return parse_num(value)


def _extract_pdf_amount(text, patterns):
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            return _parse_pdf_amount(m.group(1))
    return 0.0


def _extract_pdf_packages(text):
    patterns = [
        r'[*\-]?\s*KAP\s+ADET[İI]\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'[*\-]?\s*KAP\s+SAYISI\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'[*\-]?\s*KAP\s+ADEDI\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'[*\-]?\s*KAP\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'\bPACKAGES?\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
        r'\bCOLL[Iİ]\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
    ]
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return ''


def parse_pdf(pdf_bytes):
    """
    PDF'ten navlun, sigorta, kur ve kap bilgisini çıkarır.
    Dönen dict: {'navlun': float, 'sigorta': float, 'kur': float, 'kap': str}
    Navlun ve sigorta her zaman TRY cinsindendir.
    """
    result = {'navlun': 0.0, 'sigorta': 0.0, 'kur': 0.0, 'kap': ''}
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            # Son 2 sayfaya bak — bilgiler genellikle orada
            page_texts = [
                _normalize_pdf_text(page.extract_text() or '')
                for page in pdf.pages[-2:]
            ]
            text = ' '.join(t for t in page_texts if t).strip()
            if not text:
                return result

            result['navlun'] = _extract_pdf_amount(text, [
                r'\bNAVLUN\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
                r'\bFREIGHT\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
            ])
            result['sigorta'] = _extract_pdf_amount(text, [
                r'S[İI]G(?:ORTA)?\.?\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
                r'\bINSURANCE\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
            ])
            # Kur: "* KUR BİLGİSİ: TRY 52,5814" formatı
            result['kur'] = _extract_pdf_amount(text, [
                r'[*\-]?\s*KUR\s+B[İI]LG[İI]S[İI]\s*[:.]?\s*(?:TRY|EUR|USD)?\s*([\d.,]+)',
            ])
            result['kap'] = _extract_pdf_packages(text)
    except Exception:
        pass
    return result


def sku_grupla(df):
    """
    SKU bazında gruplandırma — INV ve PL için tekil satır üretir.
    Miktar toplanır, diğer kolonlar ilk değeri alır.
    """
    agg_dict = {col: 'first' for col in df.columns if col != 'SKU'}
    agg_dict['Miktar'] = 'sum'
    return df.groupby('SKU', sort=False).agg(agg_dict).reset_index()


def set_print(ws, print_area):
    """Yazdırma alanı ve sayfa düzenini ayarla."""
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.worksheet.properties import PageSetupProperties

    ws.print_area = print_area
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.75, bottom=0.75,
        header=0.3, footer=0.3)
    ws.print_title_rows = '1:2'
