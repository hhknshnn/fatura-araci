"""
TRY bazlı INV + PL üretim motoru.

Ülkeler: RS (Sırbistan), KZ (Kazakistan), BA (Bosna),
          GE (Gürcistan), RU (Rusya)

Tüm ülkelerde Excel Fiyat kolonu TRY cinsindendir.
Freight/Insurance: RS, KZ, BA, GE → VAR (PDF'ten TRY)
                   RU              → YOK
"""
import io

import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .constants import (
    DARK_BLUE, GOLD, TL_FMT, TRY_FMT,
    RS_INV_COLS, RS_PL_COLS,
    BA_INV_COLS, BA_PL_COLS,
    GE_INV_COLS, GE_PL_COLS,
    KZ_INV_COLS, KZ_PL_COLS,
    RU_INV_COLS, RU_PL_COLS,
)
from .helpers  import hdr, dat, parse_num, sku_grupla, set_print, brd
from .weights  import calculate_weights, get_net_list, generate_master_excel
from .templates import (
    TEMPLATE_FINDER,
    apply_ba_header, apply_ge_header,
    apply_kz_header, apply_ru_header,
    build_rs_header,
)

DS = 9   # Data start satırı (kolon başlığı bu satırda)


# ── Ortak yardımcı ────────────────────────────────────────────────────────────

def _temizle_template(wb):
    """Şablondaki eski data satırlarını sil."""
    for ws in [wb['INV'], wb['PL']]:
        if ws.max_row > DS:
            ws.delete_rows(DS + 1, ws.max_row - DS)


def _footer_try(ws_inv, last_inv, inv_total_col,
                inv_fmt, grand_total_label,
                freight_value=0.0, insurance_value=0.0,
                has_freight=True,
                label_col=None):
    """
    INV footer satırlarını yazar.
    has_freight=True  → TOTAL + FREIGHT + INSURANCE + GRAND TOTAL
    has_freight=False → TOTAL + GRAND TOTAL
    label_col: TOTAL/GRAND TOTAL etiketinin yazılacağı kolon (default: inv_total_col - 1)
    """
    tc  = get_column_letter(inv_total_col)
    lc  = label_col if label_col else inv_total_col - 1  # etiket kolonu
    vc  = inv_total_col                                   # değer kolonu

    def _cell(ws, r, col, val, bold=True, color='FFFFFF', bg=DARK_BLUE,
              align='center', fmt=None, size=10):
        c           = ws.cell(row=r, column=col, value=val)
        c.font      = Font(name='Arial', bold=bold, color=color, size=size)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border    = brd()
        if fmt:
            c.number_format = fmt
        return c

    if has_freight:
        tr, fr, ir, gr = (last_inv + 1, last_inv + 2,
                          last_inv + 3, last_inv + 4)
        for r, h in [(tr, 22), (fr, 22), (ir, 22), (gr, 28)]:
            ws_inv.row_dimensions[r].height = h

        _cell(ws_inv, tr, lc, 'TOTAL', align='center')
        _cell(ws_inv, tr, vc, f'=SUM({tc}{DS+1}:{tc}{last_inv})',
              align='right', fmt=inv_fmt)

        _cell(ws_inv, fr, lc, 'FREIGHT',   bold=True, color='000000', bg='FFFFFF', align='center')
        _cell(ws_inv, fr, vc, freight_value, bold=False, color='000000', bg='FFFFFF',
              align='right', fmt=inv_fmt)

        _cell(ws_inv, ir, lc, 'INSURANCE', bold=True, color='000000', bg='FFFFFF', align='center')
        _cell(ws_inv, ir, vc, insurance_value, bold=False, color='000000', bg='FFFFFF',
              align='right', fmt=inv_fmt)

        _cell(ws_inv, gr, lc, grand_total_label, bg=GOLD, align='center', size=11)
        _cell(ws_inv, gr, vc, f'={get_column_letter(vc)}{tr}+'
                               f'{get_column_letter(vc)}{fr}+'
                               f'{get_column_letter(vc)}{ir}',
              bg=GOLD, align='right', fmt=inv_fmt, size=11)
        return gr
    else:
        tr, gr = last_inv + 1, last_inv + 2
        ws_inv.row_dimensions[tr].height = 22
        ws_inv.row_dimensions[gr].height = 28

        _cell(ws_inv, tr, lc, 'TOTAL', align='center')
        _cell(ws_inv, tr, vc, f'=SUM({tc}{DS+1}:{tc}{last_inv})',
              align='right', fmt=inv_fmt)

        _cell(ws_inv, gr, lc, grand_total_label, bg=GOLD, align='center', size=11)
        _cell(ws_inv, gr, vc, f'={get_column_letter(vc)}{tr}',
              bg=GOLD, align='right', fmt=inv_fmt, size=11)
        return gr


def _footer_pl(ws_pl, last_pl, gross_col, net_col, label_col):
    """PL TOTAL KG footer satırı."""
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    # Boş hücreler
    for col_idx in range(1, label_col):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    def _gold(ws, r, col, val, fmt=None):
        c           = ws.cell(row=r, column=col, value=val)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        c.fill      = PatternFill('solid', fgColor=GOLD)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border    = brd()
        if fmt:
            c.number_format = fmt
        return c

    _gold(ws_pl, pl_gr, label_col, 'TOTAL KG:')

    cl = get_column_letter(gross_col)
    _gold(ws_pl, pl_gr, gross_col,
          f'=SUM({cl}{DS+1}:{cl}{last_pl})', fmt='#,##0.00')

    cl = get_column_letter(net_col)
    _gold(ws_pl, pl_gr, net_col,
          f'=SUM({cl}{DS+1}:{cl}{last_pl})', fmt='#,##0.00')

    return pl_gr


# ── Sırbistan ─────────────────────────────────────────────────────────────────

def generate_rs(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Sırbistan INV + PL — TRY, Freight/Insurance VAR."""
    # Birim ve kod temizleme
    df['Birim Cinsi (1)'] = df['Birim Cinsi (1)'].apply(
        lambda x: 'PCS' if str(x).strip() == 'AD' else x)
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_orig, _ = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()
    musteri       = str(df['Müşteri Firma Adı'].iloc[0]).strip()
    musteri_adres = 'Gospodar Jovanova 73, Belgrade'

    if pdf_fields is None:
        pdf_fields = {'navlun': 0.0, 'sigorta': 0.0, 'kap': ''}

    brut_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)[0]
    net_list  = get_net_list(brut_list, hedef_net, depo_tipi)

    wb   = Workbook()
    ws_inv = wb.active
    ws_inv.title = 'INV'

    # Header
    build_rs_header(ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date,
                    musteri, musteri_adres, len(RS_INV_COLS),
                    logo_bytes, pdf_fields)

    # Kolon başlıkları
    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(RS_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # Data satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(RS_INV_COLS):
            cn = c_idx + 1
            if src_col == '__CALC__':
                dat(ws_inv, er, cn,
                    round(parse_num(row.get('Miktar', 0)) * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=TL_FMT)
            elif out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col == 'UNIT PRICE':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt=TL_FMT)
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_inv = DS + len(df)

    # RS footer — merge A:G için özel işlem
    tr, fr, ir, gr = last_inv+1, last_inv+2, last_inv+3, last_inv+4
    for r in [tr, fr, ir, gr]:
        ws_inv.row_dimensions[r].height = 22
    ws_inv.row_dimensions[gr].height = 28

    H, I = 8, 9
    ws_inv.merge_cells(f'A{tr}:G{tr}')
    c = ws_inv.cell(row=tr, column=H, value='TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=tr, column=I, value=f'=SUM(I{DS+1}:I{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    c.fill = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TL_FMT
    c.border = brd()

    ws_inv.merge_cells(f'A{fr}:G{fr}')
    dat(ws_inv, fr, H, 'FREIGHT',   bold=True, align='center')
    dat(ws_inv, fr, I, float(pdf_fields.get('navlun', 0)), fmt=TL_FMT, align='right')

    ws_inv.merge_cells(f'A{ir}:G{ir}')
    dat(ws_inv, ir, H, 'INSURANCE', bold=True, align='center')
    dat(ws_inv, ir, I, float(pdf_fields.get('sigorta', 0)), fmt=TL_FMT, align='right')

    ws_inv.merge_cells(f'A{gr}:G{gr}')
    c = ws_inv.cell(row=gr, column=H, value='GRAND TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    c = ws_inv.cell(row=gr, column=I, value=f'=I{tr}+I{fr}+I{ir}')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = TL_FMT
    c.border = brd()

    set_print(ws_inv, f'A1:I{gr}')

    # PL sheet
    ws_pl = wb.create_sheet('PL')
    pl_widths = {'A':16,'B':14,'C':14,'D':18,'E':33,'F':6,'G':7,'H':14,'I':14}
    for col, w in pl_widths.items():
        ws_pl.column_dimensions[col].width = w

    build_rs_header(ws_pl, 'PACKING LIST', fatura_no, fatura_date,
                    musteri, musteri_adres, len(RS_PL_COLS),
                    logo_bytes, pdf_fields)

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(RS_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(RS_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)), bg=bg, align='right', fmt='#,##0')
            elif out_col == 'MASTER ITEM CODE':
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr   = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 5):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')
    ws_pl.merge_cells(f'E{pl_gr}:F{pl_gr}')
    c = ws_pl.cell(row=pl_gr, column=5, value='GRAND TOTAL')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()
    for cn, fmt in [(7, '#,##0'), (8, '#,##0.00'), (9, '#,##0.00')]:
        cl = get_column_letter(cn)
        c  = ws_pl.cell(row=pl_gr, column=cn,
                        value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor=GOLD)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.number_format = fmt
        c.border = brd()

    set_print(ws_pl, f'A1:I{pl_gr}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_orig,
                                        get_net_list(brut_orig, hedef_net, depo_tipi, hedef_brut),
                                        hedef_net=hedef_net, depo_tipi=depo_tipi)
    return buf.getvalue(), fatura_no, master_out


# ── Ortak KZ/RU/UZ motoru ────────────────────────────────────────────────────

def _generate_kz_like(ulke_kodu, df, grup_kilolari, hedef_brut, exception_skus,
                       logo_bytes, pdf_fields, hedef_net, depo_tipi, df_original,
                       inv_cols, pl_cols, inv_fmt, grand_total_label, has_freight,
                       apply_header_fn):
    """KZ, RU, UZ için ortak TRY üretim motoru."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_orig, _ = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    brut_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)[0]
    net_list  = get_net_list(brut_list, hedef_net, depo_tipi)

    freight_value   = float((pdf_fields or {}).get('navlun',  0) or 0)
    insurance_value = float((pdf_fields or {}).get('sigorta', 0) or 0)
    packages_str    = str((pdf_fields or {}).get('kap', '') or '')

    wb = openpyxl.load_workbook(TEMPLATE_FINDER[ulke_kodu]())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    apply_header_fn(ws_inv, 'COMMERCIAL INVOICE  / СЧЕТ-ФАКТУРА',
                    fatura_no, fatura_date, packages_str)
    apply_header_fn(ws_pl,  'PACKING LIST  / ТОВАРНАЯ НАКЛАДНАЯ',
                    fatura_no, fatura_date, packages_str)

    # INV başlıkları
    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(inv_cols):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # INV data satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(inv_cols):
            cn = c_idx + 1
            if src_col == '__CALC__':
                dat(ws_inv, er, cn,
                    round(parse_num(row.get('Miktar', 0)) * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=inv_fmt)
            elif out_col == 'UNIT':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col == 'UNIT PRICE':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt=inv_fmt)
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                # RU alanı boşsa ALT GRUBU -RU'dan al
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_inv, er, cn, val, bg=bg, align='left')

    last_inv = DS + len(df)
    # KZ/RU/UZ: TOTAL AMOUNT = kolon 12 (L), label = kolon 11 (K)
    # Kolon sayısı 16 olsa da TOTAL AMOUNT her zaman 12. kolonda
    INV_TOTAL_COL = 12
    gr = _footer_try(
        ws_inv, last_inv, INV_TOTAL_COL,
        inv_fmt, grand_total_label,
        freight_value, insurance_value,
        has_freight=has_freight,
        label_col=11,
    )
    set_print(ws_inv, f'A1:P{gr}')

    # PL başlıkları
    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(pl_cols):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # PL data satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(pl_cols):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'UNIT':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_pl, er, cn, val, bg=bg, align='left')

    last_pl = DS + len(df)
    # KZ/RU/UZ PL: GROSS=11(K), NET=12(L), label=10
    pl_gr = _footer_pl(ws_pl, last_pl, 11, 12, label_col=10)
    set_print(ws_pl, f'A1:L{pl_gr}')

    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    master_out = generate_master_excel(df_for_master, brut_orig,
                                        get_net_list(brut_orig, hedef_net, depo_tipi, hedef_brut),
                                        hedef_net=hedef_net, depo_tipi=depo_tipi)
    return buf.getvalue(), fatura_no, master_out


def generate_kz(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Kazakistan — TRY, Freight/Insurance VAR."""
    from .price_list import generate_price_list_pdf_kz  # opsiyonel

    excel_out, fatura_no, master_out = _generate_kz_like(
        'kz', df, grup_kilolari, hedef_brut, exception_skus,
        logo_bytes, pdf_fields, hedef_net, depo_tipi, df_original,
        KZ_INV_COLS, KZ_PL_COLS, TRY_FMT, 'GRAND TOTAL TRY',
        has_freight=True, apply_header_fn=apply_kz_header,
    )

    # Price List PDF üret (Kazakistan'a özel)
    price_list_out = None
    try:
        df_grouped = sku_grupla(df.copy())
        pdf_rows = []
        for _, row in df_grouped.iterrows():
            name_ru = row.get('Ürün Açıklaması RU', '')
            if not name_ru or str(name_ru).strip() == '':
                name_ru = row.get('ALT GRUBU -RU', '')
            pdf_rows.append({
                'sku':     str(row.get('SKU', '') or ''),
                'name_en': row.get('Ürün Açıklaması EN', ''),
                'name_ru': name_ru,
                'price':   parse_num(row.get('Fiyat', 0)),
            })
        price_list_out = generate_price_list_pdf_kz(pdf_rows, fatura_no)
    except Exception as e:
        print(f'Price List PDF üretim hatası: {e}')

    return excel_out, fatura_no, master_out, price_list_out


def generate_ru(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Rusya — TRY, Freight/Insurance YOK."""
    return _generate_kz_like(
        'ru', df, grup_kilolari, hedef_brut, exception_skus,
        logo_bytes, pdf_fields, hedef_net, depo_tipi, df_original,
        RU_INV_COLS, RU_PL_COLS, TRY_FMT, 'GRAND TOTAL TRY',
        has_freight=False, apply_header_fn=apply_ru_header,
    )


# ── Bosna ─────────────────────────────────────────────────────────────────────

def generate_ba(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Bosna — TRY, Freight/Insurance VAR."""
    df['Birim Cinsi (1)'] = df['Birim Cinsi (1)'].apply(
        lambda x: 'PCS' if str(x).strip() == 'AD' else x)
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_orig, _ = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    # Bosna'da Net Tutar (D) de toplanır
    agg_dict = {col: 'first' for col in df.columns if col != 'SKU'}
    agg_dict['Miktar'] = 'sum'
    if 'Net Tutar (D)' in df.columns:
        agg_dict['Net Tutar (D)'] = 'sum'
    df = df.groupby('SKU', sort=False).agg(agg_dict).reset_index()

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    musteri       = 'Madame Coco BH d.o.o.'
    musteri_adres = 'Ulica Vrbanja Br. 1 (SCC)-Saraybosna, Saraybosna Centar'
    destination   = 'Bosnia and Herzegovina'
    incoterm      = 'CIP'
    packages_str  = str((pdf_fields or {}).get('kap', '') or '')

    brut_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)[0]
    net_list  = get_net_list(brut_list, hedef_net, depo_tipi)

    wb = openpyxl.load_workbook(TEMPLATE_FINDER['ba']())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    apply_ba_header(ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date,
                    musteri, musteri_adres, destination, incoterm, packages_str)
    apply_ba_header(ws_pl,  'PACKING LIST',       fatura_no, fatura_date,
                    musteri, musteri_adres, destination, incoterm, packages_str)

    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(BA_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(BA_INV_COLS):
            cn = c_idx + 1
            if out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col in ('UNIT PRICE', 'TOTAL AMOUNT TRY'):
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0.00 "TRY"')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_inv = DS + len(df)
    gr_row   = last_inv + 1
    ws_inv.row_dimensions[gr_row].height = 28

    INV_UNIT_COL  = 7
    INV_TOTAL_COL = 8
    tc = get_column_letter(INV_TOTAL_COL)

    c = ws_inv.cell(row=gr_row, column=INV_UNIT_COL, value='GRAND TOTAL TRY')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = brd()

    c = ws_inv.cell(row=gr_row, column=INV_TOTAL_COL,
                    value=f'=SUM({tc}{DS+1}:{tc}{last_inv})')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = '#,##0.00 "TRY"'
    c.border = brd()

    set_print(ws_inv, f'A1:H{gr_row}')

    # PL
    for col, w in [('A',16),('B',14),('C',13),('D',18.18),('E',33),
                   ('F',6),('G',15.54),('H',20.91)]:
        ws_pl.column_dimensions[col].width = w

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(BA_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(BA_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col == 'MASTER ITEM CODE':
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr   = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 6):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    ws_pl.column_dimensions['F'].width = 12
    c = ws_pl.cell(row=pl_gr, column=6, value='TOTAL KG:')
    c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill('solid', fgColor=GOLD)
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.border = brd()

    for cn, fmt in [(7, '#,##0.00'), (8, '#,##0.00')]:
        cl = get_column_letter(cn)
        c  = ws_pl.cell(row=pl_gr, column=cn,
                        value=f'=SUM({cl}{DS+1}:{cl}{last_pl})')
        c.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill('solid', fgColor=GOLD)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.number_format = fmt
        c.border = brd()

    set_print(ws_pl, f'A1:H{pl_gr}')
    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_orig,
                                        get_net_list(brut_orig, hedef_net, depo_tipi, hedef_brut),
                                        hedef_net=hedef_net, depo_tipi=depo_tipi)
    return buf.getvalue(), fatura_no, master_out


# ── Gürcistan ─────────────────────────────────────────────────────────────────

def generate_ge(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest', df_original=None):
    """Gürcistan — TRY, Freight/Insurance VAR."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_orig, _ = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    brut_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)[0]
    net_list  = get_net_list(brut_list, hedef_net, depo_tipi)

    packages_str    = str((pdf_fields or {}).get('kap', '') or '')
    freight_value   = parse_num((pdf_fields or {}).get('navlun',  0))
    insurance_value = parse_num((pdf_fields or {}).get('sigorta', 0))

    wb = openpyxl.load_workbook(TEMPLATE_FINDER['ge']())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    apply_ge_header(ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date, packages_str)
    apply_ge_header(ws_pl,  'PACKING LIST',       fatura_no, fatura_date, packages_str)

    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(GE_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(GE_INV_COLS):
            cn = c_idx + 1
            if out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif src_col == '__CALC__':
                dat(ws_inv, er, cn,
                    round(parse_num(row.get('Miktar', 0)) * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=TRY_FMT)
            elif out_col == 'UNIT PRICE':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt=TRY_FMT)
            elif out_col in ('MASTER ITEM CODE', 'HS CODE', 'BARCODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_inv = DS + len(df)
    gr = _footer_try(
        ws_inv, last_inv, inv_total_col=8,
        inv_fmt=TRY_FMT, grand_total_label='GRAND TOTAL TRY',
        freight_value=freight_value, insurance_value=insurance_value,
        has_freight=True, label_col=7,
    )
    set_print(ws_inv, f'A1:L{gr}')

    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(GE_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(GE_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr   = _footer_pl(ws_pl, last_pl, gross_col=7, net_col=8, label_col=6)
    set_print(ws_pl, f'A1:H{pl_gr}')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    master_out = generate_master_excel(df_for_master, brut_orig,
                                        get_net_list(brut_orig, hedef_net, depo_tipi, hedef_brut),
                                        hedef_net=hedef_net, depo_tipi=depo_tipi)
    return buf.getvalue(), fatura_no, master_out
