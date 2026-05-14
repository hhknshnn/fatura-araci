import io
import os

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

from .constants import DARK_BLUE, MID_BLUE, LIGHT_BLUE, LIGHT_GRAY
from .helpers   import brd


# ── Şablon yolu bulma ─────────────────────────────────────────────────────────

def _find_template(filename):
    """Şablon dosyasını templates/ klasöründe arar."""
    current_dir = os.path.dirname(os.path.abspath(__file__))
    api_dir     = os.path.dirname(current_dir)          # api/
    root_dir    = os.path.dirname(api_dir)              # proje kökü

    candidates = [
        os.path.join(root_dir, 'templates', filename),
        os.path.join(api_dir,  'templates', filename),
        os.path.join(root_dir, filename),
        os.path.join(os.getcwd(), filename),
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    raise FileNotFoundError(f'{filename} bulunamadı. Aranan: {candidates}')


def find_rs_template_path(): return _find_template('ref_rs.xlsx')
def find_ba_template_path(): return _find_template('ref_ba.xlsx')
def find_ge_template_path(): return _find_template('ref_ge.xlsx')
def find_kz_template_path(): return _find_template('ref_kz.xlsx')
def find_ru_template_path(): return _find_template('ref_ru.xlsx')
def find_uz_template_path(): return _find_template('ref_uz.xlsx')
def find_ko_template_path(): return _find_template('ref_ko.xlsx')   # Kosova
def find_mk_template_path(): return _find_template('ref_mk.xlsx')
def find_be_template_path(): return _find_template('ref_be.xlsx')
def find_de_template_path(): return _find_template('ref_de.xlsx')
def find_nl_template_path(): return _find_template('ref_nl.xlsx')
def find_cy_template_path(): return _find_template('ref_cy.xlsx')
def find_iq_template_path(): return _find_template('ref_iq.xlsx')
def find_ly_template_path(): return _find_template('ref_ly.xlsx')
def find_lr_template_path(): return _find_template('ref_lr.xlsx')
def find_lb_template_path(): return _find_template('ref_lb.xlsx')


# ── Template path map — ülke kodu → fonksiyon ────────────────────────────────
TEMPLATE_FINDER = {
    'rs': find_rs_template_path,
    'ba': find_ba_template_path,
    'ge': find_ge_template_path,
    'kz': find_kz_template_path,
    'ru': find_ru_template_path,
    'uz': find_uz_template_path,
    'xk': find_ko_template_path,
    'mk': find_mk_template_path,
    'be': find_be_template_path,
    'de': find_de_template_path,
    'nl': find_nl_template_path,
    'cy': find_cy_template_path,
    'iq': find_iq_template_path,
    'ly': find_ly_template_path,
    'lr': find_lr_template_path,
    'lb': find_lb_template_path,
}


# ── Header uygulama fonksiyonları ─────────────────────────────────────────────
# Her şablon için header hücrelerini doldurur.
# RS — build_header ile dinamik üretilir, diğerleri şablon hücrelerine yazar.

def apply_ba_header(ws, sheet_title, fatura_no, fatura_date,
                    musteri, musteri_adres, destination, incoterm, packages=''):
    ws['A2'] = sheet_title
    ws['H3'] = str(fatura_date)
    ws['H4'] = fatura_no
    ws['H5'] = packages
    ws['H6'] = destination
    ws['H7'] = 'TURKEY'
    ws['H8'] = incoterm
    ws['A7'] = f'{musteri}\n{musteri_adres}'


def apply_ge_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['H3'] = str(fatura_date)
    ws['H4'] = fatura_no
    ws['H5'] = packages


def apply_kz_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['L3'] = str(fatura_date)
    ws['L4'] = fatura_no
    ws['L5'] = packages


def apply_ru_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['L3'] = str(fatura_date)
    ws['L4'] = fatura_no
    ws['L5'] = packages


def apply_uz_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['L3'] = str(fatura_date)
    ws['L4'] = fatura_no
    ws['L5'] = packages


def apply_ko_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    """Kosova (xk) şablonu."""
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages


def apply_mk_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages


def apply_be_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages


def apply_de_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages


def apply_nl_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    ws['A2'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages


def apply_genel_header(ws, sheet_title, fatura_no, fatura_date, packages=''):
    """USD ülkeleri (IQ, LY, LR, LB) ortak header."""
    ws['A1'] = sheet_title
    ws['I3'] = str(fatura_date)
    ws['I4'] = fatura_no
    ws['I5'] = packages


def apply_cy_header(ws, fatura_nos, fatura_date, toplam_kap):
    """Kıbrıs PL header."""
    ws['H3'] = str(fatura_date)
    ws['H4'] = fatura_nos
    ws['H5'] = toplam_kap


# ── Header uygulama map — ülke kodu → fonksiyon ───────────────────────────────
HEADER_APPLIER = {
    'ba': apply_ba_header,
    'ge': apply_ge_header,
    'kz': apply_kz_header,
    'ru': apply_ru_header,
    'uz': apply_uz_header,
    'xk': apply_ko_header,
    'mk': apply_mk_header,
    'be': apply_be_header,
    'de': apply_de_header,
    'nl': apply_nl_header,
    'iq': apply_genel_header,
    'ly': apply_genel_header,
    'lr': apply_genel_header,
    'lb': apply_genel_header,
}


# ── Sırbistan dinamik header (şablon yok, sıfırdan üretilir) ─────────────────

def build_rs_header(ws, sheet_title, fatura_no, fatura_date,
                    musteri, musteri_adres, col_count,
                    logo_bytes=None, pdf_fields=None,
                    destination='SERBIA', incoterm='CIP',
                    info_start_col=None):
    """Sırbistan INV/PL için dinamik header üretir."""
    from openpyxl.utils import column_index_from_string, get_column_letter

    last_col = get_column_letter(col_count)
    col_widths = {
        'A': 16, 'B': 14, 'C': 14, 'D': 18, 'E': 33,
        'F': 6,  'G': 7,  'H': 26, 'I': 22, 'J': 13,
        'K': 12, 'L': 16, 'M': 9,  'N': 9,  'O': 10,
    }
    for col, w in col_widths.items():
        if column_index_from_string(col) <= col_count:
            ws.column_dimensions[col].width = w

    if info_start_col is None:
        info_start_col = min(8, max(1, col_count - 1))
    info_value_col = min(col_count, info_start_col + 1)
    title_end_col  = get_column_letter(info_value_col)
    hc             = get_column_letter(max(1, info_start_col - 1))

    ws.row_dimensions[1].height = 27.0
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'].fill = PatternFill('solid', fgColor='FFFFFF')

    # Logo
    if logo_bytes:
        try:
            from PIL import Image as PILImage
            import numpy as np
            pil_img = PILImage.open(io.BytesIO(logo_bytes))
            arr  = np.array(pil_img.convert('RGB'))
            mask = (arr < 240).any(axis=2)
            rows = np.any(mask, axis=1)
            cols = np.any(mask, axis=0)
            if rows.any() and cols.any():
                rmin, rmax = np.where(rows)[0][[0, -1]]
                cmin, cmax = np.where(cols)[0][[0, -1]]
                pad     = 10
                pil_img = pil_img.crop((
                    max(0, cmin - pad), max(0, rmin - pad),
                    min(arr.shape[1], cmax + pad), min(arr.shape[0], rmax + pad)
                ))
            logo_buf = io.BytesIO()
            pil_img.save(logo_buf, format='PNG')
            logo_buf.seek(0)
            img        = XLImage(logo_buf)
            img.width  = 240
            img.height = 22
            ws.add_image(img, 'E1')
        except Exception:
            pass

    # Başlık satırı
    ws.row_dimensions[2].height = 28.0
    ws.merge_cells(f'A2:{title_end_col}2')
    c           = ws['A2']
    c.value     = sheet_title
    c.font      = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    c.fill      = PatternFill('solid', fgColor=DARK_BLUE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    for col_idx in range(info_value_col + 1, column_index_from_string(last_col) + 1):
        ws.cell(row=2, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    for i in range(3, 9):
        ws.row_dimensions[i].height = 22

    def info_label(r, txt):
        c           = ws.cell(row=r, column=info_start_col, value=txt)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
        c.fill      = PatternFill('solid', fgColor=MID_BLUE)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border    = brd()

    def info_val(r, val, bold=False):
        c           = ws.cell(row=r, column=info_value_col, value=val)
        c.font      = Font(name='Arial', bold=bold, color='000000', size=9)
        c.fill      = PatternFill('solid', fgColor=LIGHT_BLUE)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border    = brd()

    ws.merge_cells(f'A3:{hc}3')
    from .helpers import hdr as _hdr
    _hdr(ws, 3, 1, 'EXPORTER :', bg=MID_BLUE, align='left')
    info_label(3, 'INVOICE DATE :')
    info_val(3, str(fatura_date), bold=True)

    ws.row_dimensions[4].height = 32
    ws.merge_cells(f'A4:{hc}5')
    c           = ws['A4']
    c.value     = ('DEHA MAGAZACILIK EV TEKSTILI URUNLERI SAN. VE TIC. A.S.\n'
                   'Mecidiyeköy Mah. Oğuz Sok Rönesans Biz İş Merkezi No:4/14 K:4 34387 Şişli/İstanbul')
    c.font      = Font(name='Arial', size=8, color='000000')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill      = PatternFill('solid', fgColor=LIGHT_GRAY)
    c.border    = brd()
    info_label(4, 'INVOICE NO :')
    info_val(4, fatura_no, bold=True)
    info_label(5, 'PACKAGES :')
    info_val(5, str((pdf_fields or {}).get('kap', '')) if pdf_fields else '')

    ws.merge_cells(f'A6:{hc}6')
    _hdr(ws, 6, 1, 'IMPORTER :', bg=MID_BLUE, align='left')
    info_label(6, 'DESTINATION :')
    info_val(6, destination, bold=True)

    ws.row_dimensions[7].height = 32
    ws.merge_cells(f'A7:{hc}8')
    c           = ws['A7']
    c.value     = f'{musteri}\n{musteri_adres}'
    c.font      = Font(name='Arial', size=9, color='000000')
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    c.fill      = PatternFill('solid', fgColor=LIGHT_GRAY)
    c.border    = brd()
    info_label(7, 'COUNTRY OF EXPORTATION :')
    info_val(7, 'TURKEY')
    info_label(8, 'INCOTERM :')
    info_val(8, incoterm)
