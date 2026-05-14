"""
EUR bazlı INV + PL üretim motoru.

Ülkeler: XK (Kosova), MK (Makedonya), BE (Belçika), DE (Almanya), NL (Hollanda)

Excel Fiyat kolonu TRY'dir. Fiyat / eur_kuru = EUR birim fiyat.
Freight/Insurance: PDF'ten TRY olarak gelir, / eur_kuru ile EUR'a çevrilir.
Belçika'ya ek olarak Mill Test PDF üretilir.
"""
import io

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .constants import (
    DARK_BLUE, GOLD, EUR_FMT,
    EUR_INV_COLS, EUR_PL_COLS,
)
from .helpers   import hdr, dat, parse_num, sku_grupla, set_print, brd
from .weights   import calculate_weights, get_net_list, generate_master_excel
from .templates import TEMPLATE_FINDER, HEADER_APPLIER

DS = 9   # Data start satırı


def _footer_eur(ws_inv, last_inv, freight_eur, insurance_eur):
    """EUR INV footer — TOTAL + FREIGHT + INSURANCE + GRAND TOTAL EUR."""
    tr, fr, ir, gr = last_inv+1, last_inv+2, last_inv+3, last_inv+4
    for r, h in [(tr, 22), (fr, 22), (ir, 22), (gr, 28)]:
        ws_inv.row_dimensions[r].height = h

    # Kolon 8 = etiket (G indeksi), kolon 9 = değer (H indeksi)
    G, H = 8, 9
    tc = get_column_letter(H)

    def _cell(ws, r, col, val, bold=True, color='FFFFFF', bg=DARK_BLUE,
              align='center', fmt=None, size=10):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = Font(name='Arial', bold=bold, color=color, size=size)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border    = brd()
        if fmt:
            c.number_format = fmt
        return c

    _cell(ws_inv, tr, G, 'TOTAL', align='center')
    _cell(ws_inv, tr, H, f'=SUM({tc}{DS+1}:{tc}{last_inv})',
          align='right', fmt=EUR_FMT)

    _cell(ws_inv, fr, G, 'FREIGHT',   bold=True, color='000000', bg='FFFFFF', align='center')
    _cell(ws_inv, fr, H, round(freight_eur, 2),
          bold=False, color='000000', bg='FFFFFF', align='right', fmt=EUR_FMT)

    _cell(ws_inv, ir, G, 'INSURANCE', bold=True, color='000000', bg='FFFFFF', align='center')
    _cell(ws_inv, ir, H, round(insurance_eur, 2),
          bold=False, color='000000', bg='FFFFFF', align='right', fmt=EUR_FMT)

    _cell(ws_inv, gr, G, 'GRAND TOTAL EUR', bg=GOLD, align='center', size=11)
    _cell(ws_inv, gr, H,
          f'=I{tr}+I{fr}+I{ir}',
          bg=GOLD, align='right', fmt=EUR_FMT, size=11)

    return gr


def _footer_pl_eur(ws_pl, last_pl):
    """EUR PL footer — TOTAL KG."""
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, 6):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    def _gold(ws, r, col, val, fmt=None):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        c.fill      = PatternFill('solid', fgColor=GOLD)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border    = brd()
        if fmt:
            c.number_format = fmt
        return c

    _gold(ws_pl, pl_gr, 7, 'TOTAL KG:')

    for cn, fmt in [(8, '#,##0.00'), (9, '#,##0.00')]:
        cl = get_column_letter(cn)
        _gold(ws_pl, pl_gr, cn,
              f'=SUM({cl}{DS+1}:{cl}{last_pl})', fmt=fmt)

    return pl_gr


def _generate_eur(ulke_kodu, df, grup_kilolari, hedef_brut, exception_skus,
                  logo_bytes, pdf_fields, hedef_net, depo_tipi, eur_kuru, df_original):
    """Ortak EUR üretim motoru — KO, MK, BE, DE, NL."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    if not eur_kuru or eur_kuru <= 0:
        eur_kuru = 1.0

    df_for_master = df_original if df_original is not None else df
    brut_orig, _ = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    brut_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)[0]
    net_list  = get_net_list(brut_list, hedef_net, depo_tipi)

    # Freight/Insurance: PDF'ten TRY → EUR'a çevir
    freight_eur   = float((pdf_fields or {}).get('navlun',  0) or 0) / eur_kuru
    insurance_eur = float((pdf_fields or {}).get('sigorta', 0) or 0) / eur_kuru
    packages_str  = str((pdf_fields or {}).get('kap', '') or '')

    wb = openpyxl.load_workbook(TEMPLATE_FINDER[ulke_kodu]())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    apply_fn = HEADER_APPLIER[ulke_kodu]
    apply_fn(ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date, packages_str)
    apply_fn(ws_pl,  'PACKING LIST',       fatura_no, fatura_date, packages_str)

    # INV başlıkları
    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(EUR_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # INV data satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(EUR_INV_COLS):
            cn = c_idx + 1
            if src_col == '__EUR__':
                # TRY fiyat → EUR birim fiyat
                birim_eur = parse_num(row.get('Fiyat', 0)) / eur_kuru
                dat(ws_inv, er, cn, birim_eur, bg=bg, align='right', fmt=EUR_FMT)
            elif src_col == '__EUR_CALC__':
                # miktar × EUR birim fiyat
                miktar    = parse_num(row.get('Miktar', 0))
                birim_eur = parse_num(row.get('Fiyat', 0)) / eur_kuru
                dat(ws_inv, er, cn, miktar * birim_eur, bg=bg, align='right', fmt=EUR_FMT)
            elif out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_inv = DS + len(df)
    gr = _footer_eur(ws_inv, last_inv, freight_eur, insurance_eur)
    set_print(ws_inv, f'A1:L{gr}')

    # PL başlıkları
    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(EUR_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # PL data satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(EUR_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_pl = DS + len(df)
    pl_gr   = _footer_pl_eur(ws_pl, last_pl)
    set_print(ws_pl, f'A1:I{pl_gr}')

    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    master_out = generate_master_excel(df_for_master, brut_orig,
                                        get_net_list(brut_orig, hedef_net, depo_tipi, hedef_brut),
                                        hedef_net=hedef_net, depo_tipi=depo_tipi)
    return buf.getvalue(), fatura_no, master_out, fatura_date


# ── Ülke bazlı public fonksiyonlar ───────────────────────────────────────────

def generate_xk(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                eur_kuru=1.0, df_original=None):
    """Kosova — EUR, Freight/Insurance VAR."""
    excel_out, fatura_no, master_out, _ = _generate_eur(
        'xk', df, grup_kilolari, hedef_brut, exception_skus,
        logo_bytes, pdf_fields, hedef_net, depo_tipi, eur_kuru, df_original)
    return excel_out, fatura_no, master_out


def generate_mk(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                eur_kuru=1.0, df_original=None):
    """Makedonya — EUR, Freight/Insurance VAR."""
    excel_out, fatura_no, master_out, _ = _generate_eur(
        'mk', df, grup_kilolari, hedef_brut, exception_skus,
        logo_bytes, pdf_fields, hedef_net, depo_tipi, eur_kuru, df_original)
    return excel_out, fatura_no, master_out


def generate_de(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                eur_kuru=1.0, df_original=None):
    """Almanya — EUR, Freight/Insurance VAR."""
    excel_out, fatura_no, master_out, _ = _generate_eur(
        'de', df, grup_kilolari, hedef_brut, exception_skus,
        logo_bytes, pdf_fields, hedef_net, depo_tipi, eur_kuru, df_original)
    return excel_out, fatura_no, master_out


def generate_nl(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                eur_kuru=1.0, df_original=None):
    """Hollanda — EUR, Freight/Insurance VAR."""
    excel_out, fatura_no, master_out, _ = _generate_eur(
        'nl', df, grup_kilolari, hedef_brut, exception_skus,
        logo_bytes, pdf_fields, hedef_net, depo_tipi, eur_kuru, df_original)
    return excel_out, fatura_no, master_out


def generate_be(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                eur_kuru=1.0, df_original=None):
    """Belçika — EUR, Freight/Insurance VAR + Mill Test PDF."""
    excel_out, fatura_no, master_out, fatura_date = _generate_eur(
        'be', df, grup_kilolari, hedef_brut, exception_skus,
        logo_bytes, pdf_fields, hedef_net, depo_tipi, eur_kuru, df_original)

    # Mill Test PDF — Belçika'ya özel
    mill_test_out = None
    try:
        import evrak as _evrak
        form_data = {
            'faturaNo':     fatura_no,
            'faturaTarihi': str(fatura_date),
        }
        mill_test_bytes, _ = _evrak.generate_evrak_pdf('be', 'mill_test', form_data)
        mill_test_out = mill_test_bytes
    except Exception as e:
        print(f'Mill Test PDF üretim hatası: {e}')

    return excel_out, fatura_no, master_out, mill_test_out
