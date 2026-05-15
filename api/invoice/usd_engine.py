"""
USD bazlı INV + PL üretim motoru.

Ülkeler: IQ (Irak), LY (Libya), LR (Liberya), LB (Lübnan), UZ (Özbekistan)

Excel Fiyat kolonu USD cinsindendir — doğrudan kullanılır, çevrim yok.
Freight/Insurance: YOK (tüm ülkeler).
"""
import io

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .constants import (
    DARK_BLUE, GOLD, USD_FMT,
    USD_INV_COLS, USD_PL_COLS,
    KZ_INV_COLS, KZ_PL_COLS,   # Özbekistan KZ yapısını kullanır
)
from .helpers   import hdr, dat, parse_num, sku_grupla, set_print, brd
from .weights   import calculate_weights, get_net_list, generate_master_excel
from .templates import TEMPLATE_FINDER, apply_genel_header, apply_uz_header

DS = 9   # Data start satırı


def _footer_usd(ws_inv, last_inv, inv_total_col, label_col):
    """USD INV footer — TOTAL + GRAND TOTAL USD (Freight/Insurance yok)."""
    tr, gr = last_inv + 1, last_inv + 2
    ws_inv.row_dimensions[tr].height = 22
    ws_inv.row_dimensions[gr].height = 28

    tc = get_column_letter(inv_total_col)
    vc = inv_total_col

    def _cell(ws, r, col, val, bold=True, color='FFFFFF', bg=DARK_BLUE,
              align='center', fmt=None, size=10):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = Font(name='Arial', bold=bold, color=color, size=size)
        c.fill      = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=align, vertical='center')
        c.border    = brd()
        if fmt:
            c.number_format = fmt
        return c

    _cell(ws_inv, tr, label_col, 'TOTAL', align='center')
    _cell(ws_inv, tr, vc, f'=SUM({tc}{DS+1}:{tc}{last_inv})',
          align='right', fmt=USD_FMT)

    _cell(ws_inv, gr, label_col, 'GRAND TOTAL USD', bg=GOLD, align='center', size=11)
    _cell(ws_inv, gr, vc, f'={get_column_letter(vc)}{tr}',
          bg=GOLD, align='right', fmt=USD_FMT, size=11)

    return gr


def _footer_pl_usd(ws_pl, last_pl, gross_col, net_col, label_col, blank_until):
    """USD PL footer — TOTAL KG."""
    pl_gr = last_pl + 1
    ws_pl.row_dimensions[pl_gr].height = 28

    for col_idx in range(1, blank_until):
        ws_pl.cell(row=pl_gr, column=col_idx).fill = PatternFill('solid', fgColor='FFFFFF')

    def _gold(ws, r, col, val, fmt=None):
        c = ws.cell(row=r, column=col, value=val)
        c.font      = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        c.fill      = PatternFill('solid', fgColor=GOLD)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border    = brd()
        if fmt:
            c.number_format = fmt
        return c

    _gold(ws_pl, pl_gr, label_col, 'TOTAL KG:')

    cl = get_column_letter(gross_col)
    _gold(ws_pl, pl_gr, gross_col,
          f'=SUM({cl}{DS+1}:{cl}{last_pl})', fmt='#,##0.00')

    cl = get_column_letter(net_col)
    _gold(ws_pl, pl_gr, net_col,
          f'=SUM({cl}{DS+1}:{cl}{last_pl})', fmt='#,##0.00')

    return pl_gr


# ── Genel USD motoru (IQ, LY, LR, LB) ───────────────────────────────────────

def _generate_usd_genel(ulke_kodu, df, grup_kilolari, hedef_brut, exception_skus,
                         pdf_fields, hedef_net, depo_tipi, df_original):
    """IQ, LY, LR, LB için ortak USD üretim motoru."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_orig, _ = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    brut_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)[0]
    net_list  = get_net_list(brut_list, hedef_net, depo_tipi)
    packages_str = str((pdf_fields or {}).get('kap', '') or '')

    wb = openpyxl.load_workbook(TEMPLATE_FINDER[ulke_kodu]())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    apply_genel_header(ws_inv, 'COMMERCIAL INVOICE', fatura_no, fatura_date, packages_str)
    apply_genel_header(ws_pl,  'PACKING LIST',       fatura_no, fatura_date, packages_str)

    # INV başlıkları
    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(USD_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # INV data satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(USD_INV_COLS):
            cn = c_idx + 1
            if src_col == '__USD__':
                # Excel Fiyat kolonu zaten USD
                dat(ws_inv, er, cn, parse_num(row.get('Fiyat', 0)),
                    bg=bg, align='right', fmt=USD_FMT)
            elif src_col == '__USD_CALC__':
                miktar = parse_num(row.get('Miktar', 0))
                dat(ws_inv, er, cn,
                    round(miktar * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=USD_FMT)
            elif out_col == 'QTY':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_inv, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_inv = DS + len(df)
    # INV_TOTAL_COL = 9 (TOTAL AMOUNT USD), label = 8
    gr = _footer_usd(ws_inv, last_inv, inv_total_col=9, label_col=8)
    set_print(ws_inv, f'A1:L{gr}')

    # PL başlıkları
    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(USD_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # PL data satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = 23
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(USD_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'QTY':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col in ('MASTER ITEM CODE', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                dat(ws_pl, er, cn, row.get(src_col, ''), bg=bg, align='left')

    last_pl = DS + len(df)
    # PL_GROSS_COL=8, PL_NET_COL=9, label=7, blank 1..6
    pl_gr = _footer_pl_usd(ws_pl, last_pl,
                            gross_col=8, net_col=9, label_col=7, blank_until=7)
    set_print(ws_pl, f'A1:I{pl_gr}')

    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    master_out = generate_master_excel(df_for_master, brut_orig,
                                        get_net_list(brut_orig, hedef_net, depo_tipi, hedef_brut),
                                        hedef_net=hedef_net, depo_tipi=depo_tipi)
    return buf.getvalue(), fatura_no, master_out


# ── Özbekistan motoru (KZ yapısı, USD format) ─────────────────────────────────

def _generate_uz(df, grup_kilolari, hedef_brut, exception_skus,
                 pdf_fields, hedef_net, depo_tipi, df_original):
    """Özbekistan — KZ şablon yapısı, USD para birimi, Freight yok."""
    df['GTİP'] = df['GTİP'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')
    df['Asorti Barkodu'] = df['Asorti Barkodu'].apply(
        lambda x: str(int(x)) if pd.notna(x) and str(x).strip() not in ['', 'nan'] else '')

    df_for_master = df_original if df_original is not None else df
    brut_orig, _ = calculate_weights(df_for_master, grup_kilolari, hedef_brut, exception_skus)

    df = sku_grupla(df)

    fatura_no   = str(df['E-Fatura Seri Numarası'].iloc[0]).strip()
    fatura_date = df['Fatura Tarihi'].iloc[0]
    if hasattr(fatura_date, 'date'):
        fatura_date = fatura_date.date()

    brut_list = calculate_weights(df, grup_kilolari, hedef_brut, exception_skus)[0]
    net_list  = get_net_list(brut_list, hedef_net, depo_tipi)
    packages_str = str((pdf_fields or {}).get('kap', '') or '')

    wb = openpyxl.load_workbook(TEMPLATE_FINDER['uz']())
    ws_inv = wb['INV']
    ws_pl  = wb['PL']

    if ws_inv.max_row > DS:
        ws_inv.delete_rows(DS + 1, ws_inv.max_row - DS)
    if ws_pl.max_row > DS:
        ws_pl.delete_rows(DS + 1, ws_pl.max_row - DS)

    apply_uz_header(ws_inv, 'COMMERCIAL INVOICE  / СЧЕТ-ФАКТУРА',
                    fatura_no, fatura_date, packages_str)
    apply_uz_header(ws_pl,  'PACKING LIST  / ТОВАРНАЯ НАКЛАДНАЯ',
                    fatura_no, fatura_date, packages_str)

    # INV başlıkları
    ws_inv.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(KZ_INV_COLS):
        hdr(ws_inv, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # INV data satırları — KZ yapısı, USD format
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_inv.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(KZ_INV_COLS):
            cn = c_idx + 1
            if src_col == '__CALC__':
                dat(ws_inv, er, cn,
                    round(parse_num(row.get('Miktar', 0)) * parse_num(row.get('Fiyat', 0)), 2),
                    bg=bg, align='right', fmt=USD_FMT)
            elif out_col == 'UNIT':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col == 'UNIT PRICE':
                dat(ws_inv, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt=USD_FMT)
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_inv, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_inv, er, cn, val, bg=bg, align='left')

    last_inv = DS + len(df)
    # KZ yapısında INV_TOTAL_COL = 12 (L), label = 11 (K)
    gr = _footer_usd(ws_inv, last_inv, inv_total_col=12, label_col=11)
    set_print(ws_inv, f'A1:P{gr}')

    # PL başlıkları
    ws_pl.row_dimensions[DS].height = 35
    for i, (hd, _) in enumerate(KZ_PL_COLS):
        hdr(ws_pl, DS, i + 1, hd, bg=DARK_BLUE, size=9, align='center')

    # PL data satırları
    for r_idx, (_, row) in enumerate(df.iterrows()):
        er = DS + 1 + r_idx
        ws_pl.row_dimensions[er].height = None
        bg = 'FFFFFF' if r_idx % 2 == 0 else 'EBF3FB'
        for c_idx, (out_col, src_col) in enumerate(KZ_PL_COLS):
            cn = c_idx + 1
            if src_col == '__BRUT__':
                dat(ws_pl, er, cn, round(brut_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif src_col == '__NET__':
                dat(ws_pl, er, cn, round(net_list[r_idx], 2), bg=bg, align='right', fmt='#,##0.00')
            elif out_col == 'UNIT':
                dat(ws_pl, er, cn, parse_num(row.get(src_col, 0)),
                    bg=bg, align='right', fmt='#,##0')
            elif out_col in ('Master Carton Code', 'HS CODE'):
                dat(ws_pl, er, cn, str(row.get(src_col, '') or ''), bg=bg, align='left')
            else:
                val = row.get(src_col, '')
                if src_col == 'Ürün Açıklaması RU' and (not val or str(val).strip() == ''):
                    val = row.get('ALT GRUBU -RU', '')
                dat(ws_pl, er, cn, val, bg=bg, align='left')

    last_pl = DS + len(df)
    # KZ PL: GROSS=11(K), NET=12(L), label=10, blank 1..9
    pl_gr = _footer_pl_usd(ws_pl, last_pl,
                            gross_col=11, net_col=12, label_col=10, blank_until=10)
    set_print(ws_pl, f'A1:L{pl_gr}')

    ws_inv.sheet_view.topLeftCell = 'A1'
    ws_pl.sheet_view.topLeftCell  = 'A1'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    master_out = generate_master_excel(df_for_master, brut_orig,
                                        get_net_list(brut_orig, hedef_net, depo_tipi, hedef_brut),
                                        hedef_net=hedef_net, depo_tipi=depo_tipi)
    return buf.getvalue(), fatura_no, master_out


# ── Public fonksiyonlar ───────────────────────────────────────────────────────

def generate_iq(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                usd_kuru=1.0, df_original=None):
    """Irak — USD, Freight/Insurance YOK."""
    return _generate_usd_genel('iq', df, grup_kilolari, hedef_brut, exception_skus,
                                pdf_fields, hedef_net, depo_tipi, df_original)

def generate_ly(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                usd_kuru=1.0, df_original=None):
    """Libya — USD, Freight/Insurance YOK."""
    return _generate_usd_genel('ly', df, grup_kilolari, hedef_brut, exception_skus,
                                pdf_fields, hedef_net, depo_tipi, df_original)

def generate_lr(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                usd_kuru=1.0, df_original=None):
    """Liberya — USD, Freight/Insurance YOK."""
    return _generate_usd_genel('lr', df, grup_kilolari, hedef_brut, exception_skus,
                                pdf_fields, hedef_net, depo_tipi, df_original)

def generate_lb(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                usd_kuru=1.0, df_original=None):
    """Lübnan — USD, Freight/Insurance YOK."""
    return _generate_usd_genel('lb', df, grup_kilolari, hedef_brut, exception_skus,
                                pdf_fields, hedef_net, depo_tipi, df_original)

def generate_uz(df, grup_kilolari, hedef_brut, exception_skus, logo_bytes,
                pdf_fields=None, hedef_net=0, depo_tipi='serbest',
                df_original=None):
    """Özbekistan — USD (KZ yapısı), Freight/Insurance YOK."""
    return _generate_uz(df, grup_kilolari, hedef_brut, exception_skus,
                        pdf_fields, hedef_net, depo_tipi, df_original)
