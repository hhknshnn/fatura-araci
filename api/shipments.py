# api/shipments.py
# Sevkiyat kayıtları — PostgreSQL tabanlı

import io
import time
from flask import request, jsonify, send_file
from api.db import get_conn
from api.auth import get_session_from_headers

# Ülke adı → müşteri tipi eşlemesi
ULKE_MUSTERI_TIPI = {
    'SIRBİSTAN': 'kurumsal', 'BOSNA': 'kurumsal', 'GÜRCİSTAN': 'kurumsal',
    'KOSOVA': 'kurumsal', 'MAKEDONYA': 'kurumsal', 'BELÇİKA': 'kurumsal',
    'ALMANYA': 'kurumsal', 'HOLLANDA': 'kurumsal', 'KAZAKİSTAN': 'kurumsal',
    'KIBRIS': 'franchise', 'IRAK': 'franchise', 'LİBYA': 'franchise',
    'LİBERYA': 'franchise', 'LÜBNAN': 'franchise', 'ÖZBEKİSTAN': 'franchise',
    'RUSYA': 'franchise',
}

def _musteri_tipi_from_ulke(ulke):
    return ULKE_MUSTERI_TIPI.get(str(ulke).strip().upper(), 'kurumsal')


# ── TÜM SEVKİYATLARI GETİR ───────────────────────────────────────────────────
def get_all_shipments(ulke=None, durum=None, musteri_tipi=None):
    conn = get_conn()
    cur  = conn.cursor()

    query = '''
        SELECT id, ihracat_dosya_no, fatura_no, ulke, nakliye_firmasi, plaka,
               mal_bedeli_eur, navlun_eur, sigorta_eur, eur_kuru, fatura_bedeli_eur,
               fatura_bedeli_tl, durum, yukleme_tarihi, gumruk_tarihi,
               varis_tarihi, gumrukleme_bitis, created_at,
               mal_bedeli_tl, ihracat_beyanname_tl, ihracat_beyanname_eur,
               arac_bekleme, brokerage_eur, gumruk_vergisi_eur, kdv_eur,
               toplam_maliyet_eur, musteri_tipi
        FROM shipments
        WHERE 1=1
    '''
    params = []

    if ulke:
        query += ' AND unaccent(lower(ulke)) = unaccent(lower(%s))'
        params.append(ulke)
    if durum:
        query += ' AND upper(durum) = upper(%s)'
        params.append(durum)
    if musteri_tipi:
        query += ' AND musteri_tipi = %s'
        params.append(musteri_tipi)

    query += ' ORDER BY id DESC'

    cur.execute(query, params)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    return [_row_to_dict(r) for r in rows]


# ── TEK SEVKİYAT GETİR ───────────────────────────────────────────────────────
def get_shipment(shipment_id):
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute('SELECT * FROM shipments WHERE id = %s', (shipment_id,))
    row  = cur.fetchone()
    cols = [d[0] for d in cur.description]
    cur.close()
    conn.close()
    if not row:
        return None
    return dict(zip(cols, row))


# ── SEVKİYAT OLUŞTUR ─────────────────────────────────────────────────────────
def create_shipment(data):
    fatura_no = data.get('fatura_no', '')
    if fatura_no:
        conn = get_conn()
        cur  = conn.cursor()
        cur.execute('SELECT id FROM shipments WHERE fatura_no = %s', (fatura_no,))
        if cur.fetchone():
            cur.close()
            conn.close()
            raise ValueError(f'Bu fatura no zaten kayıtlı: {fatura_no}')
        cur.close()
        conn.close()

    ulke         = data.get('ulke', '')
    musteri_tipi = data.get('musteri_tipi') or _musteri_tipi_from_ulke(ulke)

    conn = get_conn()
    cur  = conn.cursor()
    cur.execute('''
        INSERT INTO shipments (
            ihracat_dosya_no, fatura_no, ulke, nakliye_firmasi, plaka,
            fatura_bedeli_tl, mal_bedeli_eur, navlun_eur, sigorta_eur,
            eur_kuru, fatura_bedeli_eur, yukleme_tarihi, gumruk_tarihi,
            varis_tarihi, durum, musteri_tipi
        ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        RETURNING id
    ''', (
        data.get('ihracat_dosya_no', ''),
        fatura_no,
        ulke,
        data.get('nakliye_firmasi', ''),
        data.get('plaka', ''),
        data.get('fatura_bedeli_tl', 0),
        data.get('mal_bedeli_eur', 0),
        data.get('navlun_eur', 0),
        data.get('sigorta_eur', 0),
        data.get('eur_kuru', 0),
        data.get('fatura_bedeli_eur', 0),
        data.get('yukleme_tarihi') or None,
        data.get('gumruk_tarihi') or None,
        data.get('varis_tarihi') or None,
        _normalize_durum(data.get('durum', 'YOLDA')),
        musteri_tipi,
    ))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()
    return new_id


# ── SEVKİYAT GÜNCELLE ────────────────────────────────────────────────────────
def update_shipment(shipment_id, data):
    conn = get_conn()
    cur  = conn.cursor()

    toplam = (
        float(data.get('ihracat_beyanname_eur', 0) or 0) +
        float(data.get('arac_bekleme', 0) or 0) +
        float(data.get('brokerage_eur', 0) or 0) +
        float(data.get('gumruk_vergisi_eur', 0) or 0) +
        float(data.get('kdv_eur', 0) or 0)
    )

    cur.execute('''
        UPDATE shipments SET
            nakliye_firmasi      = %s,
            plaka                = %s,
            ihracat_beyanname_tl  = %s,
            ihracat_beyanname_eur = %s,
            arac_bekleme         = %s,
            brokerage_eur        = %s,
            gumruk_vergisi_eur   = %s,
            kdv_eur              = %s,
            toplam_maliyet_eur   = %s,
            varis_tarihi         = %s,
            gumrukleme_bitis     = %s,
            durum                = %s
        WHERE id = %s
    ''', (
        data.get('nakliye_firmasi', ''),
        data.get('plaka', ''),
        data.get('ihracat_beyanname_tl', 0),
        data.get('ihracat_beyanname_eur', 0),
        data.get('arac_bekleme', 0),
        data.get('brokerage_eur', 0),
        data.get('gumruk_vergisi_eur', 0),
        data.get('kdv_eur', 0),
        toplam,
        data.get('varis_tarihi') or None,
        data.get('gumrukleme_bitis') or None,
        _normalize_durum(data.get('durum', 'YOLDA')),
        shipment_id,
    ))
    conn.commit()
    cur.close()
    conn.close()


# ── DASHBOARD İSTATİSTİKLERİ ─────────────────────────────────────────────────
def get_dashboard_stats():
    conn = get_conn()
    cur  = conn.cursor()

    cur.execute('SELECT COUNT(*) FROM shipments')
    toplam = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM shipments WHERE upper(durum) = 'YOLDA'")
    yolda = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM shipments WHERE upper(durum) = 'TESLİM EDİLDİ'")
    teslim = cur.fetchone()[0]

    cur.execute('SELECT COALESCE(SUM(fatura_bedeli_eur), 0) FROM shipments')
    toplam_eur = float(cur.fetchone()[0])

    cur.execute('''
        SELECT ulke, COUNT(*) as sayi
        FROM shipments
        GROUP BY ulke
        ORDER BY sayi DESC
        LIMIT 8
    ''')
    ulkeler = [{'ulke': r[0], 'sayi': r[1]} for r in cur.fetchall()]

    cur.close()
    conn.close()

    return {
        'toplam':     toplam,
        'yolda':      yolda,
        'teslim':     teslim,
        'toplam_eur': toplam_eur,
        'ulkeler':    ulkeler,
    }


# ── MALİYET RAPORU EXPORT ────────────────────────────────────────────────────
def export_shipments(ulke=None, durum=None, depo=None, musteri_tipi=None):
    rows = get_all_shipments(ulke=ulke, durum=durum, musteri_tipi=musteri_tipi)
    if depo:
        rows = [r for r in rows if str(r.get('fatura_no', '')).startswith(depo)]
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl kurulu değil'}), 500

    if not rows:
        return jsonify({'error': 'Veri bulunamadı'}), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Maliyet Raporu'

    headers = [
        'İhracat Dosya No', 'Fatura No', 'Depo', 'Ülke', 'Müşteri Tipi',
        'Nakliye Firması', 'Plaka',
        'Fatura Bedeli TL', 'Fatura Bedeli EUR', 'Mal Bedeli EUR',
        'Navlun EUR', 'Sigorta EUR', 'EUR Kuru',
        'Yükleme Tarihi', 'Gümrük Tarihi', 'Varış Tarihi', 'Gümrükleme Bitiş',
        'İhracat Beyanname TL', 'İhracat Beyanname EUR',
        'Araç Bekleme', 'Brokerage EUR', 'Gümrük Vergisi EUR', 'KDV EUR',
        'Durum',
    ]

    header_fill = PatternFill('solid', fgColor='1F3864')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    TL_FMT  = '#,##0.00 ₺'
    EUR_FMT = '#,##0.00 €'
    NUM_FMT = '#,##0.0000'

    for row_idx, s in enumerate(rows, start=2):
        fatura_no = s.get('fatura_no', '')
        depo_val  = 'ANT' if str(fatura_no).startswith('ANT') else 'IHR'

        def c(col, val, fmt=None):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if fmt: cell.number_format = fmt
            return cell

        c(1,  s.get('ihracat_dosya_no', ''))
        c(2,  fatura_no)
        c(3,  depo_val)
        c(4,  s.get('ulke', ''))
        c(5,  s.get('musteri_tipi', ''))
        c(6,  s.get('nakliye_firmasi', ''))
        c(7,  s.get('plaka', ''))
        c(8,  float(s.get('fatura_bedeli_tl', 0) or 0),  TL_FMT)
        c(9,  float(s.get('fatura_bedeli_eur', 0) or 0), EUR_FMT)
        c(10, float(s.get('mal_bedeli_eur', 0) or 0),    EUR_FMT)
        c(11, float(s.get('navlun_eur', 0) or 0),        EUR_FMT)
        c(12, float(s.get('sigorta_eur', 0) or 0),       EUR_FMT)
        c(13, float(s.get('eur_kuru', 0) or 0),          NUM_FMT)
        c(14, s.get('yukleme_tarihi', ''))
        c(15, s.get('gumruk_tarihi', ''))
        c(16, s.get('varis_tarihi', ''))
        c(17, s.get('gumrukleme_bitis', ''))
        c(18, float(s.get('ihracat_beyanname_tl', 0) or 0),  TL_FMT)
        c(19, float(s.get('ihracat_beyanname_eur', 0) or 0), EUR_FMT)
        c(20, float(s.get('arac_bekleme', 0) or 0),          EUR_FMT)
        c(21, float(s.get('brokerage_eur', 0) or 0),         EUR_FMT)
        c(22, float(s.get('gumruk_vergisi_eur', 0) or 0),    EUR_FMT)
        c(23, float(s.get('kdv_eur', 0) or 0),               EUR_FMT)
        c(24, s.get('durum', ''))

    for col_idx in range(1, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len    = len(str(ws.cell(row=1, column=col_idx).value or ''))
        for row_idx in range(2, len(rows) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='maliyet_raporu.xlsx',
    )


# ── YARDIMCI ─────────────────────────────────────────────────────────────────
def _normalize_durum(raw):
    if not raw:
        return 'YOLDA'
    s = raw.strip().upper()
    if s in ('YOLDA', 'IN TRANSIT', 'TRANSIT'):
        return 'YOLDA'
    if s in ('TESLIM EDILDI', 'TESLİM EDİLDİ', 'DELIVERED', 'TESLIM'):
        return 'TESLİM EDİLDİ'
    if s in ('VARIŞ GÜMRÜK', 'VARIS GUMRUK', 'CUSTOMS', 'GÜMRÜKTE'):
        return 'Varış Gümrük'
    if s in ('HAZIRLANIYOR', 'HAZIRLANYOR', 'PREPARING'):
        return 'HAZIRLANIYOR'
    return raw.strip()


def _row_to_dict(row):
    return {
        'id':                    row[0],
        'ihracat_dosya_no':      row[1],
        'fatura_no':             row[2],
        'ulke':                  row[3],
        'nakliye_firmasi':       row[4],
        'plaka':                 row[5],
        'mal_bedeli_eur':        float(row[6]  or 0),
        'navlun_eur':            float(row[7]  or 0),
        'sigorta_eur':           float(row[8]  or 0),
        'eur_kuru':              float(row[9]  or 0),
        'fatura_bedeli_eur':     float(row[10] or 0),
        'fatura_bedeli_tl':      float(row[11] or 0),
        'durum':                 row[12],
        'yukleme_tarihi':        str(row[13]) if row[13] else None,
        'gumruk_tarihi':         str(row[14]) if row[14] else None,
        'varis_tarihi':          str(row[15]) if row[15] else None,
        'gumrukleme_bitis':      str(row[16]) if row[16] else None,
        'created_at':            row[17],
        'mal_bedeli_tl':         float(row[18] or 0),
        'ihracat_beyanname_tl':  float(row[19] or 0),
        'ihracat_beyanname_eur': float(row[20] or 0),
        'arac_bekleme':          float(row[21] or 0),
        'brokerage_eur':         float(row[22] or 0),
        'gumruk_vergisi_eur':    float(row[23] or 0),
        'kdv_eur':               float(row[24] or 0),
        'toplam_maliyet_eur':    float(row[25] or 0),
        'musteri_tipi':          row[26] if len(row) > 26 else 'kurumsal',
    }


# ── FLASK ROUTE FONKSİYONLARI ─────────────────────────────────────────────────
def shipments_get():
    mode         = request.args.get('mode')
    ulke         = request.args.get('ulke')
    durum        = request.args.get('durum')
    musteri_tipi = request.args.get('musteri_tipi')
    sid          = request.args.get('id')

    if mode == 'dashboard':
        return jsonify({'success': True, 'stats': get_dashboard_stats()})

    if sid:
        s = get_shipment(int(sid))
        if not s:
            return jsonify({'success': False, 'error': 'Bulunamadı'}), 404
        return jsonify({'success': True, 'shipment': s})

    shipments = get_all_shipments(ulke=ulke, durum=durum, musteri_tipi=musteri_tipi)
    return jsonify({'success': True, 'shipments': shipments})


def shipments_post():
    body = request.get_json() or {}
    try:
        new_id = create_shipment(body)
        return jsonify({'success': True, 'id': new_id})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400


def shipments_put():
    body = request.get_json() or {}
    sid  = body.get('id')
    if not sid:
        return jsonify({'success': False, 'error': 'id gerekli'}), 400
    update_shipment(int(sid), body)
    return jsonify({'success': True})


def shipments_delete():
    body = request.get_json() or {}
    sid  = body.get('id')
    if not sid:
        return jsonify({'success': False, 'error': 'id gerekli'}), 400
    conn = get_conn()
    cur  = conn.cursor()
    cur.execute('DELETE FROM shipments WHERE id = %s', (int(sid),))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({'success': True})


def shipments_export():
    ulke         = request.args.get('ulke')
    durum        = request.args.get('durum')
    depo         = request.args.get('depo')
    musteri_tipi = request.args.get('musteri_tipi')
    return export_shipments(ulke=ulke, durum=durum, depo=depo, musteri_tipi=musteri_tipi)