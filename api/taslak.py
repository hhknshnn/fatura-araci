from http.server import BaseHTTPRequestHandler
import json
import base64
import io
import os
import traceback

import openpyxl

# ── CONFIG YÜKLE ──────────────────────────────────────────────────────────────
def load_config(ulke_kodu):
    """Ülkeye göre taslak config dosyasını yükle."""
    # Vercel'de çalışma dizini /var/task, config klasörü oradan erişilebilir
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_path = os.path.join(base_dir, 'config', f'taslak_{ulke_kodu}.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)

# ── TASLAK DOLDUR ─────────────────────────────────────────────────────────────
def doldur_taslak(taslak_bytes, config, form_data, mense_data=None):
    """
    Taslak Excel'e form ve menşe verilerini yaz.
    
    form_data: {referansNo, navlun, sigorta, kap, brutKg, netKg}
    mense_data: {yabanciKg, trKg} — opsiyonel, menşe adımında gelir
    """
    wb = openpyxl.load_workbook(io.BytesIO(taslak_bytes))
    sheet_name = config.get('sheet', wb.sheetnames[0])
    ws = wb[sheet_name]

    alanlar = config.get('alanlar', {})

    # ── FORM ALANLARI ─────────────────────────────────────────────────────────
    for alan_adi, alan_cfg in alanlar.items():
        hucre = alan_cfg['hucre']
        deger = form_data.get(alan_adi)
        if deger is None:
            continue

        tip    = alan_cfg.get('tip', 'metin')
        prefix = alan_cfg.get('prefix', '')

        if tip == 'sayi':
            try:    ws[hucre] = float(str(deger).replace(',', '.'))
            except: ws[hucre] = deger
        elif tip == 'tam':
            try:    ws[hucre] = int(deger)
            except: ws[hucre] = deger
        elif tip == 'metin':
            ws[hucre] = prefix + str(deger)

    # ── MENŞE ALANLARI (opsiyonel) ────────────────────────────────────────────
    if mense_data:
        mense_alanlar = config.get('menseAlanlar', {})

        yabanci_kg = mense_data.get('yabanciKg', 0)
        tr_kg      = mense_data.get('trKg', 0)

        for alan_adi, alan_cfg in mense_alanlar.items():
            hucre  = alan_cfg['hucre']
            tip    = alan_cfg.get('tip', 'sayi')
            format_str = alan_cfg.get('format', '{deger}')

            if alan_adi == 'yabanciKg':
                deger = yabanci_kg
            elif alan_adi == 'trKg':
                deger = tr_kg
            elif alan_adi == 'yabanciMetin':
                deger = format_str.replace('{deger}', str(yabanci_kg))
                ws[hucre] = deger
                continue
            elif alan_adi == 'trMetin':
                deger = format_str.replace('{deger}', str(tr_kg))
                ws[hucre] = deger
                continue
            else:
                continue

            if tip == 'sayi':
                try:    ws[hucre] = float(deger)
                except: ws[hucre] = deger

    # ── DOSYA ADI ─────────────────────────────────────────────────────────────
    ref_no  = form_data.get('referansNo', '')
    prefix  = config['alanlar']['referansNo'].get('prefix', '')
    ulke    = config.get('dosyaAdi', 'Taslak')
    dosya_adi = f"{ulke}_{prefix}{ref_no}_taslak.xlsx"

    # ── BYTES OLARAK DÖNDÜR ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), dosya_adi

# ── VERCEL HANDLER ────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS, GET')
        self.end_headers()

    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps({'status': 'ok', 'service': 'taslak'}).encode())

    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = json.loads(self.rfile.read(length))

            # Parametreler
            ulke_kodu  = body.get('ulkeKodu', 'rs')
            taslak_b64 = body.get('taslak', '')
            form_data  = body.get('formData', {})
            mense_data = body.get('menseData', None)

            # Taslak Excel bytes
            if not taslak_b64:
                raise ValueError('Taslak Excel verisi boş geldi (taslak_b64 empty)')
            taslak_bytes = base64.b64decode(taslak_b64)

            # Config yükle
            config = load_config(ulke_kodu)

            # Doldur
            excel_out, dosya_adi = doldur_taslak(
                taslak_bytes, config, form_data, mense_data)

            result = json.dumps({
                'success':  True,
                'excel':    base64.b64encode(excel_out).decode('utf-8'),
                'dosyaAdi': dosya_adi,
            })

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(result.encode('utf-8'))

        except Exception as e:
            err = json.dumps({
                'success': False,
                'error':   str(e),
                'trace':   traceback.format_exc()
            })
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(err.encode('utf-8'))

    def log_message(self, format, *args):
        pass  # Vercel loglarını temiz tut