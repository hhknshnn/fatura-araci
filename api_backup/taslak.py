from http.server import BaseHTTPRequestHandler
import json
import base64
import io
import os
import traceback
import re
import pdfplumber
import openpyxl

def _normalize_pdf_text(text):
    return re.sub(r'\s+', ' ', (text or '').replace('\u00a0', ' ')).strip()

def _extract_pdf_amount(text, patterns):
    def _parse_pdf_amount(value):
        s = str(value).strip().replace(' ', '').replace('\u00a0', '')
        if '.' in s and ',' in s:
            if s.rfind(',') > s.rfind('.'):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            s = s.replace(',', '.')
        try:
            return float(s)
        except Exception:
            return 0.0
    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            return _parse_pdf_amount(m.group(1))
    return 0.0

def parse_pdf_fields(pdf_bytes):
    result = {'navlun': 0.0, 'sigorta': 0.0, 'kap': '', 'brutKg': 0.0, 'netKg': 0.0, 'kur': 0.0}
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            # Bilgiler son sayfalarda — sadece son 2 sayfayı oku
            page_texts = [_normalize_pdf_text(page.extract_text() or '') for page in pdf.pages[-2:]]
            text = ' '.join(part for part in page_texts if part).strip()
            if not text:
                return result
            result['navlun'] = _extract_pdf_amount(text, [
                r'\bNAVLUN\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
                r'\bFREIGHT\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
            ])
            result['sigorta'] = _extract_pdf_amount(text, [
                r'S[İI]G(?:ORTA)?\.?\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
                r'\bINSURANCE\b\s*[:.]?\s*(?:TRY|TL)?\s*([\d.,]+)',
            ])
            # Kap sayısı
            kap_patterns = [
                r'[*\-]?\s*KAP\s+ADET[İI]\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
                r'[*\-]?\s*KAP\s+SAYISI\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
                r'[*\-]?\s*KAP\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
                r'\bPACKAGES?\s*[:.]?\s*(\d+(?:\s*\([^)]*\))?)',
            ]
            for p in kap_patterns:
                m = re.search(p, text, re.IGNORECASE)
                if m:
                    result['kap'] = m.group(1).strip()
                    break
            # Kur bilgisi
            result['kur'] = _extract_pdf_amount(text, [
                r'[*\-]?\s*KUR\s+B[İI]LG[İI]S[İI]\s*[:.]?\s*(?:TRY|EUR|USD)?\s*([\d.,]+)',
            ])
            # BRÜT kilo
            result['brutKg'] = _extract_pdf_amount(text, [
                r'\bB\.KG\s*[:.]?\s*([\d.,]+)',
                r'\bBRUT\s*KG\s*[:.]?\s*([\d.,]+)',
                r'\bGROSS\s*WEIGHT\s*[:.]?\s*(?:KG)?\s*([\d.,]+)',
                r'\bBRÜT\s*(?:KG|A[ĞG]IRLIK)\s*[:.]?\s*([\d.,]+)',
            ])
            # NET kilo
            result['netKg'] = _extract_pdf_amount(text, [
                r'\bN\.KG\s*[:.]?\s*([\d.,]+)',
                r'\bNET\s*KG\s*[:.]?\s*([\d.,]+)',
                r'\bNET\s*WEIGHT\s*[:.]?\s*(?:KG)?\s*([\d.,]+)',
                r'\bNET\s*A[ĞG]IRLIK\s*[:.]?\s*([\d.,]+)',
            ])
    except Exception:
        pass
    return result
# ── CONFIG YÜKLE ──────────────────────────────────────────────────────────────
def load_config(ulke_kodu):
    """Ülkeye göre taslak config dosyasını yükle."""
    # Vercel'de çalışma dizini /var/task, config klasörü oradan erişilebilir
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    config_path = os.path.join(base_dir, 'config', f'taslak_{ulke_kodu}.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)

# ── TASLAK DOLDUR ─────────────────────────────────────────────────────────────
def doldur_kibris(taslak_bytes, config, form_data):
    """Kıbrıs özel: 3 grup dinamik doldurma."""
    wb = openpyxl.load_workbook(io.BytesIO(taslak_bytes))
    ws = wb[config.get('sheet', wb.sheetnames[0])]

    gruplar   = config.get('gruplar', {})
    dosya_cfg = config.get('dosyaNo', {})

    for grup_id, grup_cfg in gruplar.items():
        kap_val   = form_data.get(grup_id + '_kap',   '')
        brut_val  = form_data.get(grup_id + '_brutKg', '')
        net_val   = form_data.get(grup_id + '_netKg',  '')

        # Grup boşsa tüm hücreleri temizle
        if not kap_val and not brut_val:
            ws[grup_cfg['kap']]    = ''
            ws[grup_cfg['brutKg']] = ''
            ws[grup_cfg['netKg']]  = ''
        else:
            ws[grup_cfg['kap']] = kap_val
            try:    ws[grup_cfg['brutKg']] = float(str(brut_val).replace(',','.'))
            except: ws[grup_cfg['brutKg']] = brut_val
            try:    ws[grup_cfg['netKg']]  = float(str(net_val).replace(',','.'))
            except: ws[grup_cfg['netKg']]  = net_val

    # Dosya no — her 3 sütuna da yaz (A8, D8, G8)
    ref_no = str(form_data.get('referansNo', ''))
    prefix = dosya_cfg.get('prefix', '')
    for hucre in ['B8', 'E8', 'H8']:
        ws[hucre] = prefix + ref_no

    dosya_adi = f"Fatura Taslak_{config['dosyaAdi']} {prefix}{ref_no}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), dosya_adi

def doldur_taslak(taslak_bytes, config, form_data, mense_data=None):
    """
    Taslak Excel'e form ve menşe verilerini yaz.
    
    form_data: {referansNo, navlun, sigorta, kap, brutKg, netKg}
    mense_data: {yabanciKg, trKg} — opsiyonel, menşe adımında gelir
    """
    wb = openpyxl.load_workbook(io.BytesIO(taslak_bytes))
    sheet_name = config.get('sheet', wb.sheetnames[0])
    ws = wb[sheet_name]

    alanlar = config.get('alanlar', {})

    # ── FORM ALANLARI ─────────────────────────────────────────────────────────
    for alan_adi, alan_cfg in alanlar.items():
        hucre = alan_cfg['hucre']
        deger = form_data.get(alan_adi)
        if deger is None:
            continue

        tip    = alan_cfg.get('tip', 'metin')
        prefix = alan_cfg.get('prefix', '')

        if tip == 'sayi':
            try:    ws[hucre] = float(str(deger).replace(',', '.'))
            except: ws[hucre] = deger
        elif tip == 'tam':
            try:    ws[hucre] = int(deger)
            except: ws[hucre] = deger
        elif tip == 'metin':
            ws[hucre] = prefix + str(deger)

    # ── MENŞE ALANLARI (opsiyonel) ────────────────────────────────────────────
    if mense_data:
        mense_alanlar = config.get('menseAlanlar', {})

        yabanci_kg = mense_data.get('yabanciKg', 0)
        tr_kg      = mense_data.get('trKg', 0)

        for alan_adi, alan_cfg in mense_alanlar.items():
            hucre  = alan_cfg['hucre']
            tip    = alan_cfg.get('tip', 'sayi')
            format_str = alan_cfg.get('format', '{deger}')

            if alan_adi == 'yabanciKg':
                deger = yabanci_kg
            elif alan_adi == 'trKg':
                deger = tr_kg
            elif alan_adi == 'yabanciMetin':
                deger = format_str.replace('{deger}', str(yabanci_kg))
                ws[hucre] = deger
                continue
            elif alan_adi == 'trMetin':
                deger = format_str.replace('{deger}', str(tr_kg))
                ws[hucre] = deger
                continue
            else:
                continue

            if tip == 'sayi':
                try:    ws[hucre] = float(deger)
                except: ws[hucre] = deger

    # ── DOSYA ADI ─────────────────────────────────────────────────────────────
    ref_no  = form_data.get('referansNo', '')
    prefix  = config['alanlar']['referansNo'].get('prefix', '')
    ulke    = config.get('dosyaAdi', 'Taslak')
    dosya_adi = f"Fatura Taslak_{ulke} {prefix}{ref_no}.xlsx"
    # ── BYTES OLARAK DÖNDÜR ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), dosya_adi

# ── VERCEL HANDLER ────────────────────────────────────────────────────────────
class handler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS, GET')
        self.end_headers()

    def do_GET(self):
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(json.dumps({'status': 'ok', 'service': 'taslak'}).encode())



    def do_POST(self):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = json.loads(self.rfile.read(length))

            # Parametreler
            # do_POST başında, ulke_kodu'ndan önce:
            action = body.get('action', 'fill')

            if action == 'parsePdf':
                pdf_b64 = body.get('pdf', '')
                if not pdf_b64:
                    raise ValueError('PDF verisi boş')
                pdf_bytes_data = base64.b64decode(pdf_b64)
                pdf_fields = parse_pdf_fields(pdf_bytes_data)
                result = json.dumps({'success': True, 'pdfFields': pdf_fields})
                self.send_response(200)
                self.send_header('Content-Type', 'application/json')
                self.send_header('Access-Control-Allow-Origin', '*')
                self.end_headers()
                self.wfile.write(result.encode('utf-8'))
                return
            ulke_kodu  = body.get('ulkeKodu', 'rs')
            taslak_b64 = body.get('taslak', '')
            form_data  = body.get('formData', {})
            mense_data = body.get('menseData', None)

            # Taslak Excel bytes
            if not taslak_b64:
                raise ValueError('Taslak Excel verisi boş geldi (taslak_b64 empty)')
            taslak_bytes = base64.b64decode(taslak_b64)

            # Config yükle
            config = load_config(ulke_kodu)

            # Doldur - Kıbrıs özel mantık
            if config.get('tip') == 'kibris':
                excel_out, dosya_adi = doldur_kibris(taslak_bytes, config, form_data)
            else:
                excel_out, dosya_adi = doldur_taslak(
                    taslak_bytes, config, form_data, mense_data)

            result = json.dumps({
                'success':  True,
                'excel':    base64.b64encode(excel_out).decode('utf-8'),
                'dosyaAdi': dosya_adi,
            })

            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(result.encode('utf-8'))

        except Exception as e:
            err = json.dumps({
                'success': False,
                'error':   str(e),
                'trace':   traceback.format_exc()
            })
            self.send_response(500)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(err.encode('utf-8'))

    def log_message(self, format, *args):
        pass  # Vercel loglarını temiz tut